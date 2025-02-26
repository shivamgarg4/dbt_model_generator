import os
from tkinter import messagebox
from openpyxl import load_workbook
import snowflake.connector
import openpyxl.worksheet.datavalidation

class ModelMapper:
    def __init__(self):
        self.audit_columns = {
            "DATA_SRC": lambda s, t: f"'{s}.{t}'",
            "CREATE_DT": lambda *_: "CURRENT_TIMESTAMP()",
            "CREATE_BY": lambda *_: "CAST(CURRENT_USER() AS VARCHAR(200))",
            "CREATE_PGM": lambda s, t: f"'{s}.{t}'",
            "UPDATE_DT": lambda *_: "CURRENT_TIMESTAMP()",
            "UPDATE_BY": lambda *_: "CAST(CURRENT_USER() AS VARCHAR(200))",
            "UPDATE_PGM": lambda s, t: f"'{s}.{t}'"
        }

    def generate_model_mapping(self, file_path):
        """Generate model mapping using Snowflake connection"""
        try:
            # Load and validate workbook
            workbook, mapping_sheet, config_sheet = self._load_workbook(file_path)
            
            # Get table information
            target_table, source_table = self._get_table_info(mapping_sheet)
            
            # Parse table names
            source_info, target_info = self._parse_table_names(source_table, target_table)
            
            # Get Snowflake configuration
            snowflake_config = self._get_snowflake_config(config_sheet)
            
            # Connect to Snowflake and get column information
            columns = self._get_snowflake_columns(snowflake_config, source_info)
            
            # Update mapping sheet
            mapped_count, total_count = self._update_mapping_sheet(
                mapping_sheet, 
                columns, 
                source_info, 
                target_info
            )
            
            # Save workbook
            workbook.save(file_path)
            
            return mapped_count, total_count

        except Exception as e:
            raise Exception(f"Model mapping generation failed: {str(e)}")

    def _load_workbook(self, file_path):
        """Load and validate Excel workbook"""
        if not file_path:
            raise ValueError("No Excel file selected")
            
        workbook = load_workbook(file_path)
        if 'Mapping' not in workbook.sheetnames or 'Config' not in workbook.sheetnames:
            raise ValueError("Required sheets (Mapping and Config) not found in Excel")
            
        return workbook, workbook['Mapping'], workbook['Config']

    def _get_table_info(self, mapping_sheet):
        """Extract target and source table information from mapping sheet"""
        target_table = None
        source_table = None
        
        for row in range(1, mapping_sheet.max_row + 1):
            cell_value = mapping_sheet.cell(row=row, column=1).value
            if cell_value == 'TARGET_TABLE':
                target_table = mapping_sheet.cell(row=row, column=2).value
            elif cell_value == 'SOURCE_TABLE':
                source_table = mapping_sheet.cell(row=row, column=2).value
            if target_table and source_table:
                break

        if not source_table:
            raise ValueError("Source table not found in mapping sheet")
            
        return target_table, source_table

    def _parse_table_names(self, source_table, target_table):
        """Parse source and target table names"""
        try:
            source_db, source_schema, source_table_name = source_table.split('.')
            target_parts = target_table.split('.')
            target_schema = target_parts[0] if len(target_parts) > 1 else None
            target_table_name = target_parts[-1]
            
            source_info = {
                'database': source_db,
                'schema': source_schema,
                'table': source_table_name,
                'short_name': f"{source_schema}.{source_table_name}"
            }
            
            target_info = {
                'schema': target_schema,
                'table': target_table_name
            }
            
            return source_info, target_info
            
        except ValueError:
            raise ValueError("Source table should be in format: DATABASE.SCHEMA.TABLE")

    def _get_snowflake_config(self, config_sheet):
        """Extract Snowflake configuration from Config sheet"""
        snowflake_config = {}
        in_snowflake_section = False
        
        for row in range(2, config_sheet.max_row + 1):
            param = config_sheet.cell(row=row, column=1).value
            if not param:
                continue
            if param == 'Snowflake Configuration':
                in_snowflake_section = True
                continue
            if in_snowflake_section and param in ['ROLE', 'WAREHOUSE', 'DATABASE', 'ACCOUNT', 'USER', 'AUTHENTICATOR']:
                value = config_sheet.cell(row=row, column=2).value
                if value:
                    snowflake_config[param] = value
                    
        return snowflake_config

    def _get_snowflake_columns(self, config, source_info):
        """Get column information from Snowflake"""
        conn = None
        cursor = None
        try:
            conn = snowflake.connector.connect(
                account=config['ACCOUNT'],
                user=config['USER'],
                authenticator=config['AUTHENTICATOR'],
                warehouse=config['WAREHOUSE'],
                role=config['ROLE'],
                database=source_info['database'],
                schema=source_info['schema']
            )

            cursor = conn.cursor()
            cursor.execute(f"""
                SELECT COLUMN_NAME
                FROM {source_info['database']}.INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = '{source_info['schema']}'
                AND TABLE_NAME = '{source_info['table']}'
                ORDER BY ORDINAL_POSITION
            """)
            
            return cursor.fetchall()
            
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    def _update_mapping_sheet(self, mapping_sheet, columns, source_info, target_info):
        """Update mapping sheet with column mappings"""
        start_row = 5  # Column mappings start from row 5
        mapped_count = 0
        total_count = 0
        
        # Add materialization field if it doesn't exist
        materialization_exists = False
        for row in range(1, mapping_sheet.max_row + 1):
            if mapping_sheet.cell(row=row, column=1).value == 'MATERIALIZATION':
                materialization_exists = True
                break
                
        if not materialization_exists:
            # Find the row after SOURCE_NAME
            source_name_row = None
            for row in range(1, mapping_sheet.max_row + 1):
                if mapping_sheet.cell(row=row, column=1).value == 'SOURCE_NAME':
                    source_name_row = row
                    break
            
            if source_name_row:
                # Insert a new row after SOURCE_NAME
                mapping_sheet.insert_rows(source_name_row + 1)
                
                # Add MATERIALIZATION label and dropdown
                mat_label = mapping_sheet.cell(row=source_name_row + 1, column=1, value='MATERIALIZATION')
                mat_label.fill = mapping_sheet.cell(row=source_name_row, column=1).fill  # Copy fill from SOURCE_NAME
                
                # Add data validation for materialization
                mat_cell = mapping_sheet.cell(row=source_name_row + 1, column=2)
                dv = openpyxl.worksheet.datavalidation.DataValidation(
                    type="list",
                    formula1='"incremental,truncate_load"',
                    allow_blank=True
                )
                mapping_sheet.add_data_validation(dv)
                dv.add(mat_cell)
        
        for row in range(start_row, mapping_sheet.max_row + 1):
            target_column = mapping_sheet.cell(row=row, column=2).value
            if not target_column:
                continue

            total_count += 1
            target_column = target_column.upper()
            
            # Handle audit columns
            if target_column in self.audit_columns:
                mapping_sheet.cell(row=row, column=3).value = source_info['short_name']
                mapping_sheet.cell(row=row, column=4).value = self.audit_columns[target_column](
                    target_info['schema'], 
                    target_info['table']
                )
                mapped_count += 1
            else:
                # Look for matching column in source
                for col in columns:
                    if col[0].upper() == target_column:
                        mapping_sheet.cell(row=row, column=3).value = source_info['short_name']
                        mapping_sheet.cell(row=row, column=4).value = col[0]
                        mapped_count += 1
                        break
                        
        return mapped_count, total_count
