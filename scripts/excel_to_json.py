import pandas as pd
import json
import os
from openpyxl import load_workbook
from datetime import datetime
from scripts.utils import get_config_from_sheet, get_table_info_from_sheet

def convert_excel_to_json(excel_path, json_output_path=None):
    """
    Convert Excel mapping to JSON configuration
    
    Args:
        excel_path (str): Path to Excel file
        json_output_path (str, optional): Path to save JSON file. If None, generates default path
        
    Returns:
        dict: Configuration dictionary
    """
    try:
        # Load workbook
        workbook = load_workbook(excel_path)
        
        # Get required sheets
        if 'Mapping' not in workbook.sheetnames or 'Config' not in workbook.sheetnames:
            raise ValueError("Required sheets (Mapping and Config) not found in Excel")
            
        mapping_sheet = workbook['Mapping']
        config_sheet = workbook['Config']

        # Get target and source details
        target_table, source_table = get_table_info_from_sheet(mapping_sheet)
        if not target_table or not source_table:
            raise ValueError("Target or source table not found in mapping sheet")

        # Parse tables
        target_parts = target_table.split('.')
        target_schema = target_parts[0] if len(target_parts) > 1 else "RAW_CORE"
        target_table_name = target_parts[-1]

        source_parts = source_table.split('.')
        if len(source_parts) != 3:
            raise ValueError("Source table should be in format: DATABASE.SCHEMA.TABLE")
        source_db, source_schema, source_table = source_parts

        # Get Snowflake config
        snowflake_config = get_config_from_sheet(config_sheet)

        # Create configuration dictionary
        config = {
            "Connection": snowflake_config,
            "Source": {
                "Database": source_db,
                "Schema": source_schema,
                "DP View": source_table
            },
            "Target": {
                "Schema": target_schema,
                "Table Name": target_table_name,
                "MATERIALIZATION": "incremental"  # Default materialization
            },
            "DAG": {
                "DAG Type": "",
                "Cron Values": "",
                "Dependencies": []
            }
        }

        # Save JSON if path provided
        if json_output_path:
            os.makedirs(os.path.dirname(os.path.abspath(json_output_path)), exist_ok=True)
            with open(json_output_path, 'w') as f:
                json.dump(config, f, indent=4)

        return config

    except Exception as e:
        raise Exception(f"Error converting Excel to JSON: {str(e)}")

def read_excel_sheets(excel_file_path):
    """Read all sheets into a dictionary of DataFrames using openpyxl engine."""
    return pd.read_excel(excel_file_path, sheet_name=None, engine='openpyxl')

def process_sheets(sheets_dict):
    """Process each sheet and store the relevant data in a dictionary."""
    data_dict = {}
    for sheet_name, df in sheets_dict.items():
        if not df.empty:
            first_row_dict = df.iloc[0].to_dict()
            if 'Dependency Schema' in df.columns and 'Dependency Object' in df.columns:
                first_row_dict['Dependency Schema'] = df['Dependency Schema'].dropna().tolist()
                first_row_dict['Dependency Object'] = df['Dependency Object'].dropna().tolist()
            data_dict[sheet_name] = first_row_dict
        else:
            data_dict[sheet_name] = "This sheet is empty."
    return data_dict

def save_json(data_dict, json_output_path):
    """Ensure the directory exists and save the dictionary to a JSON file."""
    os.makedirs(os.path.dirname(json_output_path), exist_ok=True)
    with open(json_output_path, 'w') as json_file:
        json.dump(data_dict, json_file, indent=4)
