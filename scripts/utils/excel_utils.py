from openpyxl import load_workbook
import openpyxl

def get_config_from_sheet(config_sheet):
    """Extract Snowflake configuration from Config sheet"""
    config = {}
    in_snowflake_section = False
    for row in range(2, config_sheet.max_row + 1):
        param = config_sheet.cell(row=row, column=1).value
        if not param:
            continue
        if param == 'Snowflake Configuration':
            in_snowflake_section = True
            continue
        if in_snowflake_section and param in ['ROLE', 'WAREHOUSE', 'DATABASE', 'ACCOUNT', 'USER', 'AUTHENTICATOR']:
            value = config_sheet.cell(row=row, column=2).value
            if value:
                config[param] = value
    return config

def get_table_info_from_sheet(mapping_sheet):
    """Get target and source table information from mapping sheet"""
    target_table = None
    source_table = None
    for row in range(1, mapping_sheet.max_row + 1):
        cell_value = mapping_sheet.cell(row=row, column=1).value
        if cell_value == 'TARGET_TABLE':
            target_table = mapping_sheet.cell(row=row, column=2).value
        elif cell_value == 'SOURCE_TABLE':
            source_table = mapping_sheet.cell(row=row, column=2).value
        if target_table and source_table:
            break
    return target_table, source_table 