import json
from openpyxl import load_workbook
import snowflake.connector
import openpyxl.worksheet.datavalidation
import os
from   .model_mapper  import ModelMapper
def format_columns(columns):
    """Format columns for the dbt model."""
    try:
        max_length = max(len(column[0]) for column in columns)
        return [f"    {column[0].ljust(max_length)}" for column in columns]
    except Exception as e:
        raise Exception(f"An error occurred while formatting columns: {e}")

def replace_audit_columns(columns, source_schema_name, source_view_name, target_schema_name, target_table_name):
    """Replace audit columns with static values."""
    try:
        audit_columns = {
            "DATA_SRC": f"'{source_schema_name}.{source_view_name}' AS DATA_SRC",
            "CREATE_DT": "CURRENT_TIMESTAMP() AS CREATE_DT",
            "CREATE_BY": "CAST(CURRENT_USER() AS VARCHAR(200)) AS CREATE_BY",
            "CREATE_PGM": f"'{target_schema_name}.{target_table_name}' AS CREATE_PGM",
            "UPDATE_DT": "CURRENT_TIMESTAMP() AS UPDATE_DT",
            "UPDATE_BY": "CAST(CURRENT_USER() AS VARCHAR(200)) AS UPDATE_BY",
            "UPDATE_PGM": f"'{target_schema_name}.{target_table_name}' AS UPDATE_PGM"
        }
        return [f"    {audit_columns.get(column.split()[0].strip(), column)}" for column in columns]
    except Exception as e:
        raise Exception(f"An error occurred while replacing audit columns: {e}")

def generate_lnd_dbt_model_file(config, file_path):
    
    """Generate the dbt model file."""
    try:
        with open(config) as f:
            config = json.load(f)

        # Extract source information
        source_name = config['Source']['Name']


        workbook = load_workbook(file_path)
        mapping_sheet, config_sheet = workbook['Mapping'], workbook['Config']

        """Extract target and source table information from mapping sheet"""
        target_table = None
        source_table = None

        for row in range(1, mapping_sheet.max_row + 1):
            cell_value = mapping_sheet.cell(row=row, column=1).value
            if cell_value == 'TARGET_TABLE':
                target_table = mapping_sheet.cell(row=row, column=2).value
            elif cell_value == 'SOURCE_TABLE':
                source_table = mapping_sheet.cell(row=row, column=2).value
            if target_table and source_table:
                break

        if not source_table:
            raise ValueError("Source table not found in mapping sheet")


        source_db, source_schema, source_table_name = source_table.split('.')
        target_parts = target_table.split('.')
        target_schema = target_parts[0] if len(target_parts) > 1 else None
        target_table_name = target_parts[-1]

        """Extract Snowflake configuration from Config sheet"""
        snowflake_config = {}
        in_snowflake_section = False

        for row in range(2, config_sheet.max_row + 1):
            param = config_sheet.cell(row=row, column=1).value
            if not param:
                continue
            if param == 'Snowflake Configuration':
                in_snowflake_section = True
                continue
            if in_snowflake_section and param in ['ROLE', 'WAREHOUSE', 'DATABASE', 'ACCOUNT', 'USER', 'AUTHENTICATOR']:
                value = config_sheet.cell(row=row, column=2).value
                if value:
                    snowflake_config[param] = value

        """Get column information from Snowflake"""
        conn = None
        cursor = None
        conn = snowflake.connector.connect(
            account=snowflake_config['ACCOUNT'],
            user=snowflake_config['USER'],
            authenticator=snowflake_config['AUTHENTICATOR'],
            warehouse=snowflake_config['WAREHOUSE'],
            role=snowflake_config['ROLE'],
            database=source_db,
            schema=source_schema
        )

        print(conn)

        cursor = conn.cursor()
        cursor.execute(f"""
            SELECT COLUMN_NAME
            FROM {source_db}.INFORMATION_SCHEMA.COLUMNS
             WHERE TABLE_SCHEMA = '{source_schema}'
            AND TABLE_NAME = '{source_table_name}'
            ORDER BY ORDINAL_POSITION
        """)

        columns = cursor.fetchall()
        formatted_columns = format_columns(columns)
        replaced_columns = replace_audit_columns(formatted_columns, source_schema, source_table_name, target_schema, target_table_name)
        model_config = f"""
        {{
            config(
                schema = '{target_schema}',
                alias = '{target_table_name}',
                tags = '{target_table_name}',
                materialized='table',
                transient = false
            )
        }}
        -- Model: {target_schema}.{target_table_name}
        -- Source: {source_db}.{source_schema}.{source_table_name}
        """

        model_config += "\n\nSELECT\n"
        model_config += ",\n".join(replaced_columns) + "\n"
        model_config += f"FROM {{{{ source('{source_name}', '{source_table_name}') }}}}\n"

        # Create output directory if it doesn't exist
        output_dir = 'models'
        os.makedirs(output_dir, exist_ok=True)
    
        # Generate file name
        model_name = f"{target_schema}.{target_table_name}"
        file_name = f"{model_name}.sql"
        file_path = os.path.join(output_dir, file_name)

        # Write model file
        with open(file_path, 'w') as f:
            f.write(model_config)

        return file_path
    except Exception as e:
        raise Exception(f"An error occurred while generating the dbt model: {e}")

def get_snowflake_connection(config):
    """
    Create a Snowflake connection using the provided configuration
    
    Args:
        config (dict): Dictionary containing Snowflake connection parameters
            Required keys: account, user, password, warehouse, role, database, schema
            
    Returns:
        snowflake.connector.SnowflakeConnection: Snowflake connection object
    """
    try:
        conn = snowflake.connector.connect(
            account=config['account'],
            user=config['user'],
            password=config['password'],
            warehouse=config['warehouse'],
            role=config['role'],
            database=config['database'],
            schema=config['schema']
        )
        return conn
    except Exception as e:
        raise Exception(f"Failed to connect to Snowflake: {str(e)}")
