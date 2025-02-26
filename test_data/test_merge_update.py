import os
import sys
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Add the parent directory to the path so we can import the scripts
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import the necessary functions
from scripts.dbt_model_generator import create_dbt_model_from_json

def create_test_mapping_file(file_path, exclude_columns="CREATE_DT,CREATE_BY"):
    """Create a test mapping file with specified exclude columns"""
    # Create a new workbook
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Create the Mapping sheet
    mapping_sheet = wb.create_sheet('Mapping')
    
    # Add table information
    mapping_sheet.cell(row=1, column=1, value='TARGET_TABLE')
    mapping_sheet.cell(row=1, column=2, value='TEST_SCHEMA.TEST_TABLE')
    
    mapping_sheet.cell(row=2, column=1, value='SOURCE_TABLE')
    mapping_sheet.cell(row=2, column=2, value='SOURCE_SCHEMA.SOURCE_TABLE')
    
    mapping_sheet.cell(row=3, column=1, value='SOURCE_TYPE')
    mapping_sheet.cell(row=3, column=2, value='source')
    
    mapping_sheet.cell(row=4, column=1, value='SOURCE_NAME')
    mapping_sheet.cell(row=4, column=2, value='SOURCE_SCHEMA')
    
    mapping_sheet.cell(row=5, column=1, value='MATERIALIZATION')
    mapping_sheet.cell(row=5, column=2, value='incremental')
    
    mapping_sheet.cell(row=6, column=1, value='UNIQUE_KEY')
    mapping_sheet.cell(row=6, column=2, value='ID')
    
    mapping_sheet.cell(row=7, column=1, value='MINUS_LOGIC_REQUIRED')
    mapping_sheet.cell(row=7, column=2, value='N')
    
    mapping_sheet.cell(row=8, column=1, value='MERGE_UPDATE_EXCLUDE_COLUMNS')
    mapping_sheet.cell(row=8, column=2, value=exclude_columns)
    
    # Add column mapping headers
    mapping_sheet.cell(row=10, column=1, value='S.NO')
    mapping_sheet.cell(row=10, column=2, value='TARGET COLUMN')
    mapping_sheet.cell(row=10, column=3, value='SOURCE COLUMN')
    mapping_sheet.cell(row=10, column=4, value='LOGIC')
    
    # Apply bold formatting to headers
    for col in range(1, 5):
        mapping_sheet.cell(row=10, column=col).font = Font(bold=True)
    
    # Add sample columns
    columns = [
        (1, 'ID', 'ID', 'ID'),
        (2, 'NAME', 'NAME', 'NAME'),
        (3, 'VALUE', 'VALUE', 'VALUE'),
        (4, 'STATUS', 'STATUS', 'STATUS'),
        (5, 'DATA_SRC', '', "'SOURCE_SCHEMA.SOURCE_TABLE'"),
        (6, 'CREATE_DT', '', 'CURRENT_TIMESTAMP()'),
        (7, 'CREATE_BY', '', "CAST(CURRENT_USER() AS VARCHAR(200))"),
        (8, 'CREATE_PGM', '', "'TEST_SCHEMA.TEST_TABLE'"),
        (9, 'UPDATE_DT', '', 'CURRENT_TIMESTAMP()'),
        (10, 'UPDATE_BY', '', "CAST(CURRENT_USER() AS VARCHAR(200))"),
        (11, 'UPDATE_PGM', '', "'TEST_SCHEMA.TEST_TABLE'")
    ]
    
    # Add columns to the sheet
    for i, (sno, target, source, logic) in enumerate(columns, 11):
        mapping_sheet.cell(row=i, column=1, value=sno)
        mapping_sheet.cell(row=i, column=2, value=target)
        mapping_sheet.cell(row=i, column=3, value=source)
        mapping_sheet.cell(row=i, column=4, value=logic)
    
    # Save the workbook
    wb.save(file_path)
    
    return file_path

def create_test_json_config(file_path, mapping_file):
    """Create a test JSON configuration based on the mapping file"""
    # Load the mapping file
    wb = openpyxl.load_workbook(mapping_file)
    mapping_sheet = wb['Mapping']
    
    # Extract basic information
    target_table = mapping_sheet.cell(row=1, column=2).value
    source_table = mapping_sheet.cell(row=2, column=2).value
    source_type = mapping_sheet.cell(row=3, column=2).value
    source_name = mapping_sheet.cell(row=4, column=2).value
    materialization = mapping_sheet.cell(row=5, column=2).value
    unique_key = mapping_sheet.cell(row=6, column=2).value
    
    # Split target and source tables into schema and table name
    target_schema, target_table_name = target_table.split('.')
    source_schema, source_table_name = source_table.split('.')
    
    # Create the configuration
    config = {
        "Source": {
            "Type": source_type,
            "Database": source_schema,
            "Schema": source_schema,
            "Table Name": source_table_name,
            "Name": source_name
        },
        "Target": {
            "Schema": target_schema,
            "Table Name": target_table_name,
            "materialization": materialization
        },
        "Columns": []
    }
    
    # Add unique key if provided
    if unique_key:
        if ',' in unique_key:
            config['Target']['unique_key'] = [key.strip() for key in unique_key.split(',')]
        else:
            config['Target']['unique_key'] = unique_key
    
    # Add columns
    for row in range(11, 22):  # Rows 11-21 contain the columns
        target_col = mapping_sheet.cell(row=row, column=2).value
        source_col = mapping_sheet.cell(row=row, column=3).value
        logic = mapping_sheet.cell(row=row, column=4).value
        
        if target_col:
            config['Columns'].append({
                "Target Column": target_col,
                "Source Table": source_schema,
                "Logic": logic
            })
    
    # Write the configuration to a file
    with open(file_path, 'w') as f:
        json.dump(config, f, indent=2)
    
    return file_path

def test_merge_update_columns():
    """Test the merge_update_columns feature"""
    try:
        # Create test directories
        os.makedirs('models', exist_ok=True)
        
        print("=" * 80)
        print("TESTING MERGE_UPDATE_COLUMNS FEATURE")
        print("=" * 80)
        
        # Test with default exclude columns (CREATE_DT, CREATE_BY)
        print("\nTest 1: Default exclude columns (CREATE_DT, CREATE_BY)")
        print("-" * 50)
        
        # Create test mapping file with default exclude columns
        mapping_file_default = create_test_mapping_file('test_data/test_mapping_default.xlsx')
        print(f"Created test mapping file: {mapping_file_default}")
        
        # Create test JSON config
        json_file_default = create_test_json_config('test_data/test_config_default.json', mapping_file_default)
        print(f"Created test JSON config: {json_file_default}")
        
        # Generate DBT model
        wb_default = openpyxl.load_workbook(mapping_file_default)
        mapping_sheet_default = wb_default['Mapping']
        model_file_default = create_dbt_model_from_json(json_file_default, mapping_sheet_default)
        print(f"Generated DBT model: {model_file_default}")
        
        # Test with custom exclude columns (CREATE_DT, CREATE_BY, UPDATE_DT, UPDATE_BY)
        print("\nTest 2: Custom exclude columns (CREATE_DT, CREATE_BY, UPDATE_DT, UPDATE_BY)")
        print("-" * 50)
        
        # Create test mapping file with custom exclude columns
        mapping_file_custom = create_test_mapping_file('test_data/test_mapping_custom.xlsx', 
                                                      "CREATE_DT,CREATE_BY,UPDATE_DT,UPDATE_BY")
        print(f"Created test mapping file: {mapping_file_custom}")
        
        # Create test JSON config
        json_file_custom = create_test_json_config('test_data/test_config_custom.json', mapping_file_custom)
        print(f"Created test JSON config: {json_file_custom}")
        
        # Generate DBT model
        wb_custom = openpyxl.load_workbook(mapping_file_custom)
        mapping_sheet_custom = wb_custom['Mapping']
        model_file_custom = create_dbt_model_from_json(json_file_custom, mapping_sheet_custom)
        print(f"Generated DBT model: {model_file_custom}")
        
        # Print the contents of the generated models
        print("\n" + "=" * 80)
        print("VERIFICATION RESULTS")
        print("=" * 80)
        
        # Check the default model
        print("\nDefault model (exclude CREATE_DT, CREATE_BY):")
        print("-" * 50)
        with open(model_file_default, 'r') as f:
            content = f.read()
            print(content)
            
            # Check if merge_update_columns is included
            if "merge_update_columns" in content:
                print("\n✓ PASSED - merge_update_columns is included in the model")
                
                # Check if CREATE_DT and CREATE_BY are excluded
                if "'CREATE_DT'" not in content.split("merge_update_columns")[1].split("]")[0] and "'CREATE_BY'" not in content.split("merge_update_columns")[1].split("]")[0]:
                    print("✓ PASSED - CREATE_DT and CREATE_BY are excluded from merge_update_columns")
                else:
                    print("✗ FAILED - CREATE_DT or CREATE_BY are not excluded from merge_update_columns")
                
                # Check if ID (unique key) is excluded
                if "'ID'" not in content.split("merge_update_columns")[1].split("]")[0]:
                    print("✓ PASSED - ID (unique key) is excluded from merge_update_columns")
                else:
                    print("✗ FAILED - ID (unique key) is not excluded from merge_update_columns")
                
                # Check if expected columns are included
                update_columns_section = content.split("merge_update_columns")[1].split("]")[0]
                expected_columns = ["'NAME'", "'VALUE'", "'STATUS'", "'DATA_SRC'", "'CREATE_PGM'", "'UPDATE_PGM'"]
                all_included = all(col in update_columns_section for col in expected_columns)
                if all_included:
                    print("✓ PASSED - All expected columns are included in merge_update_columns")
                else:
                    print("✗ FAILED - Some expected columns are missing from merge_update_columns")
                    print(f"Debug - Update columns section: {update_columns_section}")
                    print(f"Debug - Expected columns: {expected_columns}")
            else:
                print("\n✗ FAILED - merge_update_columns is not included in the model")
        
        # Check the custom model
        print("\nCustom model (exclude CREATE_DT, CREATE_BY, UPDATE_DT, UPDATE_BY):")
        print("-" * 50)
        with open(model_file_custom, 'r') as f:
            content = f.read()
            print(content)
            
            # Check if merge_update_columns is included
            if "merge_update_columns" in content:
                print("\n✓ PASSED - merge_update_columns is included in the model")
                
                # Check if CREATE_DT, CREATE_BY, UPDATE_DT, and UPDATE_BY are excluded
                update_columns_section = content.split("merge_update_columns")[1].split("]")[0]
                excluded = all(col not in update_columns_section for col in 
                              ["'CREATE_DT'", "'CREATE_BY'", "'UPDATE_DT'", "'UPDATE_BY'"])
                if excluded:
                    print("✓ PASSED - CREATE_DT, CREATE_BY, UPDATE_DT, and UPDATE_BY are excluded from merge_update_columns")
                else:
                    print("✗ FAILED - Some columns are not excluded from merge_update_columns")
                    print(f"Debug - Update columns section: {update_columns_section}")
                
                # Check if ID (unique key) is excluded
                if "'ID'" not in update_columns_section:
                    print("✓ PASSED - ID (unique key) is excluded from merge_update_columns")
                else:
                    print("✗ FAILED - ID (unique key) is not excluded from merge_update_columns")
                
                # Check if expected columns are included
                expected_columns = ["'NAME'", "'VALUE'", "'STATUS'", "'DATA_SRC'", "'CREATE_PGM'", "'UPDATE_PGM'"]
                all_included = all(col in update_columns_section for col in expected_columns)
                if all_included:
                    print("✓ PASSED - All expected columns are included in merge_update_columns")
                else:
                    print("✗ FAILED - Some expected columns are missing from merge_update_columns")
                    print(f"Debug - Update columns section: {update_columns_section}")
                    print(f"Debug - Expected columns: {expected_columns}")
            else:
                print("\n✗ FAILED - merge_update_columns is not included in the model")
        
        print("\n" + "=" * 80)
        print("TEST COMPLETED")
        print("=" * 80)
    except Exception as e:
        import traceback
        print(f"ERROR: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    test_merge_update_columns() 