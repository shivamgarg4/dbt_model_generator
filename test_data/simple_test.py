import os
import sys
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Add the parent directory to the path so we can import the scripts
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import the necessary functions
from scripts.dbt_model_generator import create_dbt_model_from_json

def create_simple_json_config():
    """Create a simple JSON configuration file for testing"""
    config = {
        "Source": {
            "Type": "source",
            "Database": "SOURCE_SCHEMA",
            "Schema": "SOURCE_SCHEMA",
            "Table Name": "SOURCE_TABLE",
            "Name": "SOURCE_SCHEMA"
        },
        "Target": {
            "Schema": "TEST_SCHEMA",
            "Table Name": "TEST_TABLE",
            "materialization": "incremental",
            "unique_key": "ID"
        },
        "Columns": [
            {
                "Target Column": "ID",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "ID"
            },
            {
                "Target Column": "NAME",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "NAME"
            },
            {
                "Target Column": "VALUE",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "VALUE"
            },
            {
                "Target Column": "STATUS",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "STATUS"
            },
            {
                "Target Column": "DATA_SRC",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "'SOURCE_SCHEMA.SOURCE_TABLE'"
            },
            {
                "Target Column": "CREATE_DT",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "CURRENT_TIMESTAMP()"
            },
            {
                "Target Column": "CREATE_BY",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "CAST(CURRENT_USER() AS VARCHAR(200))"
            },
            {
                "Target Column": "CREATE_PGM",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "'TEST_SCHEMA.TEST_TABLE'"
            },
            {
                "Target Column": "UPDATE_DT",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "CURRENT_TIMESTAMP()"
            },
            {
                "Target Column": "UPDATE_BY",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "CAST(CURRENT_USER() AS VARCHAR(200))"
            },
            {
                "Target Column": "UPDATE_PGM",
                "Source Table": "SOURCE_SCHEMA",
                "Logic": "'TEST_SCHEMA.TEST_TABLE'"
            }
        ]
    }
    
    # Write the configuration to a file
    with open('test_data/simple_config.json', 'w') as f:
        json.dump(config, f, indent=2)
    
    return 'test_data/simple_config.json'

def create_simple_mapping_sheet(minus_logic_required='Y'):
    """Create a simple mapping sheet for testing"""
    # Create a new workbook
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Create the Mapping sheet
    mapping_sheet = wb.create_sheet('Mapping')
    
    # Add table information
    mapping_sheet.cell(row=1, column=1, value='TARGET_TABLE')
    mapping_sheet.cell(row=1, column=2, value='TEST_SCHEMA.TEST_TABLE')
    
    mapping_sheet.cell(row=2, column=1, value='SOURCE_TABLE')
    mapping_sheet.cell(row=2, column=2, value='SOURCE_SCHEMA.SOURCE_TABLE')
    
    mapping_sheet.cell(row=3, column=1, value='SOURCE_TYPE')
    mapping_sheet.cell(row=3, column=2, value='source')
    
    mapping_sheet.cell(row=4, column=1, value='SOURCE_NAME')
    mapping_sheet.cell(row=4, column=2, value='SOURCE_SCHEMA')
    
    mapping_sheet.cell(row=5, column=1, value='MATERIALIZATION')
    mapping_sheet.cell(row=5, column=2, value='incremental')
    
    mapping_sheet.cell(row=6, column=1, value='UNIQUE_KEY')
    mapping_sheet.cell(row=6, column=2, value='ID')
    
    mapping_sheet.cell(row=7, column=1, value='MINUS_LOGIC_REQUIRED')
    mapping_sheet.cell(row=7, column=2, value=minus_logic_required)
    
    # Save the workbook
    wb.save('test_data/simple_mapping.xlsx')
    
    return wb['Mapping']

def run_simple_test():
    """Run a simple test of the dbt_model_generator functionality"""
    try:
        print("Creating test directories...")
        os.makedirs('models', exist_ok=True)
        
        print("Creating simple JSON config...")
        config_file = create_simple_json_config()
        
        print("Creating simple mapping sheet...")
        mapping_sheet = create_simple_mapping_sheet('Y')
        
        print("Generating DBT model...")
        model_file = create_dbt_model_from_json(config_file, mapping_sheet)
        
        print(f"Generated model file: {model_file}")
        
        print("\nContents of the generated model:")
        print("-" * 50)
        with open(model_file, 'r') as f:
            print(f.read())
            
        print("\nTest completed successfully!")
    except Exception as e:
        import traceback
        print(f"ERROR: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    run_simple_test() 