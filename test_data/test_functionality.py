import os
import sys
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# Add the parent directory to the path so we can import the scripts
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import the necessary functions
from scripts.dbt_model_generator import create_dbt_model_from_json
from dag_generator_app import DAGGeneratorApp

def create_test_mapping_file(file_path, minus_logic_required='N'):
    """Create a test mapping file"""
    # Create a new workbook
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Create the Mapping sheet
    mapping_sheet = wb.create_sheet('Mapping')
    
    # Add header
    mapping_sheet.merge_cells('A1:D1')
    header_cell = mapping_sheet.cell(row=1, column=1, value="TEST_TABLE Mapping")
    header_cell.font = Font(bold=True, color='FFFFFF')
    header_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_cell.alignment = Alignment(horizontal='center')
    
    # Add table information
    target_label = mapping_sheet.cell(row=2, column=1, value='TARGET_TABLE')
    target_label.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    mapping_sheet.cell(row=2, column=2, value='TEST_SCHEMA.TEST_TABLE')
    
    source_label = mapping_sheet.cell(row=3, column=1, value='SOURCE_TABLE')
    source_label.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    mapping_sheet.cell(row=3, column=2, value='SOURCE_SCHEMA.SOURCE_TABLE')
    
    source_type_label = mapping_sheet.cell(row=4, column=1, value='SOURCE_TYPE')
    source_type_label.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    mapping_sheet.cell(row=4, column=2, value='source')
    
    source_name_label = mapping_sheet.cell(row=5, column=1, value='SOURCE_NAME')
    source_name_label.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    mapping_sheet.cell(row=5, column=2, value='SOURCE_SCHEMA')
    
    materialization_label = mapping_sheet.cell(row=6, column=1, value='MATERIALIZATION')
    materialization_label.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    mapping_sheet.cell(row=6, column=2, value='incremental')
    
    unique_key_label = mapping_sheet.cell(row=7, column=1, value='UNIQUE_KEY')
    unique_key_label.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    mapping_sheet.cell(row=7, column=2, value='ID')
    
    minus_logic_label = mapping_sheet.cell(row=8, column=1, value='MINUS_LOGIC_REQUIRED')
    minus_logic_label.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    mapping_sheet.cell(row=8, column=2, value=minus_logic_required)
    
    # Add blank row
    mapping_sheet.append([])
    
    # Add column headers
    headers = ['S.NO', 'TargetColumn', 'Source Table', 'Logic/Mapping/Constant Value']
    header_row = 10
    for col, header in enumerate(headers, start=1):
        cell = mapping_sheet.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Add columns
    columns = [
        ('ID', 'SOURCE_SCHEMA', 'ID'),
        ('NAME', 'SOURCE_SCHEMA', 'NAME'),
        ('VALUE', 'SOURCE_SCHEMA', 'VALUE'),
        ('STATUS', 'SOURCE_SCHEMA', 'STATUS'),
        ('DATA_SRC', 'SOURCE_SCHEMA', "'SOURCE_SCHEMA.SOURCE_TABLE'"),
        ('CREATE_DT', 'SOURCE_SCHEMA', 'CURRENT_TIMESTAMP()'),
        ('CREATE_BY', 'SOURCE_SCHEMA', "CAST(CURRENT_USER() AS VARCHAR(200))"),
        ('CREATE_PGM', 'SOURCE_SCHEMA', "'TEST_SCHEMA.TEST_TABLE'"),
        ('UPDATE_DT', 'SOURCE_SCHEMA', 'CURRENT_TIMESTAMP()'),
        ('UPDATE_BY', 'SOURCE_SCHEMA', "CAST(CURRENT_USER() AS VARCHAR(200))"),
        ('UPDATE_PGM', 'SOURCE_SCHEMA', "'TEST_SCHEMA.TEST_TABLE'")
    ]
    
    for i, (col, src, logic) in enumerate(columns, start=1):
        row = header_row + i
        mapping_sheet.cell(row=row, column=1, value=i)
        mapping_sheet.cell(row=row, column=2, value=col)
        mapping_sheet.cell(row=row, column=3, value=src)
        mapping_sheet.cell(row=row, column=4, value=logic)
    
    # Add JOIN_TABLES section
    join_row = header_row + len(columns) + 2
    join_cell = mapping_sheet.cell(row=join_row, column=1, value='JOIN_TABLES')
    join_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Add WHERE_CONDITIONS section
    where_row = join_row + 5
    where_cell = mapping_sheet.cell(row=where_row, column=1, value='WHERE_CONDITIONS')
    where_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Add GROUP BY section
    group_row = where_row + 2
    group_cell = mapping_sheet.cell(row=group_row, column=1, value='GROUP BY')
    group_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Create Config sheet
    config_sheet = wb.create_sheet('Config')
    
    # Save the workbook
    wb.save(file_path)
    
    return file_path

def create_test_json_config(file_path, mapping_file):
    """Create a test JSON configuration file based on the mapping file"""
    # Load the mapping file
    wb = openpyxl.load_workbook(mapping_file)
    mapping_sheet = wb['Mapping']
    
    # Extract information from the mapping file
    target_table = None
    source_table = None
    source_type = None
    source_name = None
    materialization = None
    unique_key = None
    
    for row in range(1, 10):
        label = mapping_sheet.cell(row=row, column=1).value
        value = mapping_sheet.cell(row=row, column=2).value
        
        if label == 'TARGET_TABLE':
            target_table = value
        elif label == 'SOURCE_TABLE':
            source_table = value
        elif label == 'SOURCE_TYPE':
            source_type = value
        elif label == 'SOURCE_NAME':
            source_name = value
        elif label == 'MATERIALIZATION':
            materialization = value
        elif label == 'UNIQUE_KEY':
            unique_key = value
    
    # Parse target and source tables
    target_parts = target_table.split('.')
    source_parts = source_table.split('.')
    
    # Extract column mappings
    columns = []
    header_row = 10  # Assuming header row is 10
    
    for row in range(header_row + 1, mapping_sheet.max_row + 1):
        target_col = mapping_sheet.cell(row=row, column=2).value
        source_table_col = mapping_sheet.cell(row=row, column=3).value
        logic = mapping_sheet.cell(row=row, column=4).value
        
        if not target_col:
            # Stop when we reach an empty row or the JOIN_TABLES section
            next_row_value = mapping_sheet.cell(row=row+1, column=1).value
            if next_row_value and next_row_value == 'JOIN_TABLES':
                break
            continue
        
        columns.append({
            'Target Column': target_col,
            'Source Table': source_table_col,
            'Logic': logic or target_col  # Use target column as default logic
        })
    
    # Create the configuration
    config = {
        'Source': {
            'Type': source_type,
            'Database': source_parts[0],
            'Schema': source_parts[0],
            'Table Name': source_parts[1],
            'Name': source_name
        },
        'Target': {
            'Schema': target_parts[0],
            'Table Name': target_parts[1],
            'materialization': materialization
        },
        'Columns': columns
    }
    
    # Add unique key if provided
    if unique_key:
        if ',' in unique_key:
            config['Target']['unique_key'] = [key.strip() for key in unique_key.split(',')]
        else:
            config['Target']['unique_key'] = unique_key
    
    # Write the configuration to a file
    with open(file_path, 'w') as f:
        json.dump(config, f, indent=2)
    
    return file_path

def test_functionality():
    """Test the new functionality"""
    try:
        # Create test directories
        os.makedirs('models', exist_ok=True)
        
        print("=" * 80)
        print("TESTING DBT MODEL GENERATOR WITH NEW FUNCTIONALITY")
        print("=" * 80)
        
        print("\nTesting with MINUS_LOGIC_REQUIRED = Y")
        print("-" * 50)
        # Create test mapping file with MINUS_LOGIC_REQUIRED = Y
        mapping_file_y = create_test_mapping_file('test_data/test_mapping_y.xlsx', 'Y')
        print(f"Created test mapping file: {mapping_file_y}")
        
        # Create test JSON config
        json_file_y = create_test_json_config('test_data/test_config_y.json', mapping_file_y)
        print(f"Created test JSON config: {json_file_y}")
        
        # Generate DBT model
        print("Loading workbook...")
        wb_y = openpyxl.load_workbook(mapping_file_y)
        print("Getting mapping sheet...")
        mapping_sheet_y = wb_y['Mapping']
        print("Calling create_dbt_model_from_json...")
        model_file_y = create_dbt_model_from_json(json_file_y, mapping_sheet_y)
        print(f"Generated DBT model: {model_file_y}")
        
        # Print the contents of the generated models
        print("\n" + "=" * 80)
        print("VERIFICATION RESULTS")
        print("=" * 80)
        
        # Check if the source alias is used
        with open(model_file_y, 'r') as f:
            content = f.read()
            print("\nSource alias test:")
            if "source" in content and "AS source" in content:
                print("✓ PASSED - Source alias 'source' is used in the model")
            else:
                print("✗ FAILED - Source alias 'source' is not used in the model")
        
        # Check if minus logic is included
        with open(model_file_y, 'r') as f:
            content = f.read()
            print("\nMinus logic test:")
            if "MINUS" in content:
                print("✓ PASSED - Minus logic is included in the model")
                
                # Check if minus logic is implemented as a subquery
                if "FROM\n(" in content and "MINUS" in content and ")" in content:
                    print("✓ PASSED - Minus logic is implemented as a subquery")
                else:
                    print("✗ FAILED - Minus logic is not implemented as a subquery")
                    
                # Check if audit columns are outside the subquery
                audit_columns = ["DATA_SRC", "CREATE_DT", "CREATE_BY", "CREATE_PGM", 
                               "UPDATE_DT", "UPDATE_BY", "UPDATE_PGM"]
                
                audit_outside = True
                for col in audit_columns:
                    if f"{col}" not in content.split("FROM\n(")[0]:
                        audit_outside = False
                        break
                        
                if audit_outside:
                    print("✓ PASSED - Audit columns are outside the subquery")
                else:
                    print("✗ FAILED - Audit columns are not outside the subquery")
            else:
                print("✗ FAILED - Minus logic is not included in the model")
        
        print("\nContents of model with minus logic:")
        print("-" * 50)
        with open(model_file_y, 'r') as f:
            print(f.read())
        
        print("\n" + "=" * 80)
        print("TEST COMPLETED")
        print("=" * 80)
    except Exception as e:
        import traceback
        print(f"ERROR: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    test_functionality() 