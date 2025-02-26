#!/usr/bin/env python
"""
Setup Scripts for DBT Model Generator
-------------------------------------
This script checks for and creates any missing script files needed by the application.
"""

import os
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='setup.log'
)

# Required script files with their content
REQUIRED_SCRIPTS = {
    'scripts/excel_to_json.py': '''
import json
import openpyxl

def convert_excel_to_json(excel_file, json_file):
    """Convert Excel mapping file to JSON configuration"""
    workbook = openpyxl.load_workbook(excel_file)
    mapping_sheet = workbook['Mapping']
    
    # Extract table information
    config = {}
    for row in range(1, 10):
        key = mapping_sheet.cell(row=row, column=1).value
        if not key:
            continue
        value = mapping_sheet.cell(row=row, column=2).value
        if key == 'TARGET_TABLE':
            config['target_table'] = value
        elif key == 'SOURCE_TABLE':
            config['source_table'] = value
        elif key == 'SOURCE_TYPE':
            config['source_type'] = value
        elif key == 'SOURCE_NAME':
            config['source_name'] = value
        elif key == 'MATERIALIZATION':
            config['materialization'] = value
    
    # Write to JSON file
    with open(json_file, 'w') as f:
        json.dump(config, f, indent=2)
    
    return json_file
''',
    'scripts/dbt_model_generator.py': '''
import json
import os

def create_dbt_model_from_json(json_file, mapping_sheet=None):
    """Create a DBT model file from JSON configuration"""
    with open(json_file, 'r') as f:
        config = json.load(f)
    
    # Extract target information
    target_schema = config['Target']['Schema']
    target_table = config['Target']['Table Name']
    
    # Create model file path
    model_dir = os.path.join('models', target_schema)
    os.makedirs(model_dir, exist_ok=True)
    model_file = os.path.join(model_dir, f"{target_table}.sql")
    
    # Generate model content
    model_content = generate_model_content(config)
    
    # Write model file
    with open(model_file, 'w') as f:
        f.write(model_content)
    
    return model_file

def generate_model_content(config):
    """Generate the SQL content for the DBT model"""
    # This is a simplified version - the actual implementation would be more complex
    model_content = f"-- DBT model for {config['Target']['Schema']}.{config['Target']['Table Name']}\\n\\n"
    
    # Add config block
    model_content += "{{ config(\\n"
    model_content += f"    materialized='{config['Target']['materialization']}',\\n"
    if 'unique_key' in config['Target']:
        unique_keys = config['Target']['unique_key']
        if isinstance(unique_keys, list) and len(unique_keys) == 1:
            model_content += f"    unique_key='{unique_keys[0]}',\\n"
        elif isinstance(unique_keys, list) and len(unique_keys) > 1:
            model_content += f"    unique_key=['{\"', '\".join(unique_keys)}'],\\n"
    model_content += ") }}\\n\\n"
    
    # Add select statement
    model_content += "SELECT\\n"
    
    # Add columns
    for i, column in enumerate(config['Columns']):
        comma = "," if i < len(config['Columns']) - 1 else ""
        model_content += f"    {column['Logic']} AS {column['Target Column']}{comma}\\n"
    
    # Add from clause
    source_type = config['Source']['Type']
    if source_type.lower() == 'source':
        model_content += f"FROM {{{{ source('{config['Source']['Name']}', '{config['Source']['Table Name']}') }}}}\\n"
    else:
        model_content += f"FROM {{{{ ref('{config['Source']['Table Name']}') }}}}\\n"
    
    return model_content
''',
    'scripts/dbt_job_generator.py': '''
import json
import os

def create_dbt_job_file(json_file, job_output_path):
    """Create a DBT job file from JSON configuration"""
    with open(json_file, 'r') as f:
        config = json.load(f)
    
    # Extract target information
    target_schema = config['Target']['Schema']
    target_table = config['Target']['Table Name']
    
    # Create job file path
    os.makedirs(job_output_path, exist_ok=True)
    job_file = os.path.join(job_output_path, f"{target_schema}_{target_table}.yml")
    
    # Generate job content
    job_content = generate_job_content(config)
    
    # Write job file
    with open(job_file, 'w') as f:
        f.write(job_content)
    
    return job_file

def generate_job_content(config):
    """Generate the YAML content for the DBT job"""
    # This is a simplified version - the actual implementation would be more complex
    job_content = f"# DBT job for {config['Target']['Schema']}.{config['Target']['Table Name']}\\n\\n"
    
    job_content += "jobs:\\n"
    job_content += f"  - name: {config['Target']['Schema']}_{config['Target']['Table Name']}\\n"
    job_content += "    steps:\\n"
    job_content += f"      - dbt run --select {config['Target']['Schema']}.{config['Target']['Table Name']}\\n"
    
    return job_content
''',
    'scripts/dag_generators.py': '''
import json
import os

def create_dataset_dependency_dag(json_file, dag_file_path):
    """Create a dataset dependency DAG file"""
    with open(json_file, 'r') as f:
        config = json.load(f)
    
    # Extract target information
    target_schema = config['Target']['Schema']
    target_table = config['Target']['Table Name']
    
    # Generate DAG content
    dag_content = generate_dataset_dag_content(config, target_schema, target_table)
    
    # Write DAG file
    with open(dag_file_path, 'w') as f:
        f.write(dag_content)
    
    return dag_file_path

def create_cron_dag(json_file, dag_file_path):
    """Create a CRON scheduled DAG file"""
    with open(json_file, 'r') as f:
        config = json.load(f)
    
    # Extract target information
    target_schema = config['Target']['Schema']
    target_table = config['Target']['Table Name']
    
    # Get schedule from config or use default
    schedule = config['DAG'].get('Schedule', '0 */4 * * *')
    
    # Generate DAG content
    dag_content = generate_cron_dag_content(config, target_schema, target_table, schedule)
    
    # Write DAG file
    with open(dag_file_path, 'w') as f:
        f.write(dag_content)
    
    return dag_file_path

def create_sns_dag(json_file, dag_file_path):
    """Create an SNS triggered DAG file"""
    with open(json_file, 'r') as f:
        config = json.load(f)
    
    # Extract target information
    target_schema = config['Target']['Schema']
    target_table = config['Target']['Table Name']
    
    # Generate DAG content
    dag_content = generate_sns_dag_content(config, target_schema, target_table)
    
    # Write DAG file
    with open(dag_file_path, 'w') as f:
        f.write(dag_content)
    
    return dag_file_path

def generate_dataset_dag_content(config, schema, table):
    """Generate content for a dataset dependency DAG"""
    dag_id = f"{schema}_{table}_dag"
    
    content = f"""
from airflow import DAG
from airflow.operators.bash import BashOperator
from datetime import datetime, timedelta

default_args = {{
    'owner': 'airflow',
    'depends_on_past': False,
    'email_on_failure': False,
    'email_on_retry': False,
    'retries': 1,
    'retry_delay': timedelta(minutes=5),
}}

dag = DAG(
    '{dag_id}',
    default_args=default_args,
    description='Dataset dependency DAG for {schema}.{table}',
    schedule_interval=None,
    start_date=datetime(2023, 1, 1),
    catchup=False,
)

run_dbt = BashOperator(
    task_id='run_dbt_model',
    bash_command='dbt run --select {schema}.{table}',
    dag=dag,
)
"""
    return content

def generate_cron_dag_content(config, schema, table, schedule):
    """Generate content for a CRON scheduled DAG"""
    dag_id = f"{schema}_{table}_dag"
    
    content = f"""
from airflow import DAG
from airflow.operators.bash import BashOperator
from datetime import datetime, timedelta

default_args = {{
    'owner': 'airflow',
    'depends_on_past': False,
    'email_on_failure': False,
    'email_on_retry': False,
    'retries': 1,
    'retry_delay': timedelta(minutes=5),
}}

dag = DAG(
    '{dag_id}',
    default_args=default_args,
    description='CRON scheduled DAG for {schema}.{table}',
    schedule_interval='{schedule}',
    start_date=datetime(2023, 1, 1),
    catchup=False,
)

run_dbt = BashOperator(
    task_id='run_dbt_model',
    bash_command='dbt run --select {schema}.{table}',
    dag=dag,
)
"""
    return content

def generate_sns_dag_content(config, schema, table):
    """Generate content for an SNS triggered DAG"""
    dag_id = f"{schema}_{table}_sns_dag"
    
    content = f"""
from airflow import DAG
from airflow.operators.bash import BashOperator
from airflow.sensors.external_task import ExternalTaskSensor
from datetime import datetime, timedelta

default_args = {{
    'owner': 'airflow',
    'depends_on_past': False,
    'email_on_failure': False,
    'email_on_retry': False,
    'retries': 1,
    'retry_delay': timedelta(minutes=5),
}}

dag = DAG(
    '{dag_id}',
    default_args=default_args,
    description='SNS triggered DAG for {schema}.{table}',
    schedule_interval=None,
    start_date=datetime(2023, 1, 1),
    catchup=False,
)

run_dbt = BashOperator(
    task_id='run_dbt_model',
    bash_command='dbt run --select {schema}.{table}',
    dag=dag,
)
"""
    return content
''',
    'scripts/utils.py': '''
import os
import json
import logging

def ensure_directory_exists(directory_path):
    """Ensure that a directory exists, creating it if necessary"""
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)
        logging.info(f"Created directory: {directory_path}")
    return directory_path

def load_json_file(file_path):
    """Load and parse a JSON file"""
    try:
        with open(file_path, 'r') as f:
            return json.load(f)
    except Exception as e:
        logging.error(f"Error loading JSON file {file_path}: {str(e)}")
        return None

def save_json_file(data, file_path):
    """Save data to a JSON file"""
    try:
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=2)
        logging.info(f"Saved JSON file: {file_path}")
        return True
    except Exception as e:
        logging.error(f"Error saving JSON file {file_path}: {str(e)}")
        return False

def get_file_extension(file_path):
    """Get the extension of a file"""
    _, ext = os.path.splitext(file_path)
    return ext.lower()
''',
    'scripts/model_mapper.py': '''
import json
import os
import logging

def map_columns(source_columns, target_columns, mapping_rules=None):
    """Map source columns to target columns based on rules"""
    if mapping_rules is None:
        mapping_rules = {}
    
    column_mapping = {}
    
    # Apply explicit mapping rules first
    for source_col, target_col in mapping_rules.items():
        if source_col in source_columns:
            column_mapping[source_col] = target_col
    
    # For remaining columns, try to find matches by name
    for source_col in source_columns:
        if source_col in column_mapping:
            continue
        
        # Check if column exists in target with same name
        if source_col in target_columns:
            column_mapping[source_col] = source_col
        # Check if lowercase version exists
        elif source_col.lower() in [col.lower() for col in target_columns]:
            matching_col = next(col for col in target_columns if col.lower() == source_col.lower())
            column_mapping[source_col] = matching_col
    
    return column_mapping

def generate_mapping_file(source_table, target_table, column_mapping, output_path):
    """Generate a mapping file for source to target table"""
    mapping = {
        "Source": {
            "Table": source_table,
            "Columns": list(column_mapping.keys())
        },
        "Target": {
            "Table": target_table,
            "Columns": list(column_mapping.values())
        },
        "Mapping": column_mapping
    }
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Write mapping to file
    with open(output_path, 'w') as f:
        json.dump(mapping, f, indent=2)
    
    logging.info(f"Generated mapping file: {output_path}")
    return output_path
'''
}

def check_script_files():
    """Check if required script files exist and create them if they don't"""
    logging.info("Checking for required script files...")
    
    # Ensure scripts directory exists
    if not os.path.exists('scripts'):
        os.makedirs('scripts')
        logging.info("Created scripts directory")
    
    # Check each required script
    for script_path, script_content in REQUIRED_SCRIPTS.items():
        if not os.path.exists(script_path):
            # Create directory if needed
            script_dir = os.path.dirname(script_path)
            if script_dir and not os.path.exists(script_dir):
                os.makedirs(script_dir)
                logging.info(f"Created directory: {script_dir}")
            
            # Create script file
            with open(script_path, 'w') as f:
                f.write(script_content)
            logging.info(f"Created script file: {script_path}")
        else:
            logging.info(f"Script file already exists: {script_path}")

if __name__ == "__main__":
    logging.info("Starting setup_scripts.py")
    check_script_files()
    logging.info("Setup complete") 