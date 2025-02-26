import os
import json
import openpyxl
from openpyxl.styles import Font
from scripts.dbt_model_generator import create_dbt_model_from_json

def create_test_mapping_file(file_path):
    """Create a test mapping file with the issues we need to fix"""
    # Create directory if it doesn't exist
    if os.path.dirname(file_path):
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    mapping_sheet = wb.active
    mapping_sheet.title = "Mapping"
    
    # Add headers and configuration
    mapping_sheet.cell(row=1, column=1, value='TARGET_TABLE')
    mapping_sheet.cell(row=1, column=2, value='DW.D_OPCO')
    
    mapping_sheet.cell(row=2, column=1, value='SOURCE_TABLE')
    mapping_sheet.cell(row=2, column=2, value='EBI_DEV_DB.LND_CORE.D_DP_OPCO')
    
    mapping_sheet.cell(row=3, column=1, value='SOURCE_TYPE')
    mapping_sheet.cell(row=3, column=2, value='source')
    
    mapping_sheet.cell(row=4, column=1, value='SOURCE_NAME')
    mapping_sheet.cell(row=4, column=2, value='LND_CORE')
    
    mapping_sheet.cell(row=5, column=1, value='MATERIALIZATION')
    mapping_sheet.cell(row=5, column=2, value='incremental')
    
    mapping_sheet.cell(row=6, column=1, value='UNIQUE_KEY')
    mapping_sheet.cell(row=6, column=2, value='OPCO_CD')
    
    mapping_sheet.cell(row=7, column=1, value='MINUS_LOGIC_REQUIRED')
    mapping_sheet.cell(row=7, column=2, value='Y')
    
    mapping_sheet.cell(row=8, column=1, value='MERGE_UPDATE_EXCLUDE_COLUMNS')
    mapping_sheet.cell(row=8, column=2, value='CREATE_DT,CREATE_BY,CREATE_PGM')
    
    # Add column headers
    mapping_sheet.cell(row=10, column=1, value='S.NO')
    mapping_sheet.cell(row=10, column=2, value='TargetColumn')
    mapping_sheet.cell(row=10, column=3, value='Source Table')
    mapping_sheet.cell(row=10, column=4, value='Logic/Mapping/Constant Value')
    
    # Apply bold formatting to headers
    for col in range(1, 5):
        mapping_sheet.cell(row=10, column=col).font = Font(bold=True)
    
    # Add sample columns
    columns = [
        (1, 'OPCO_ID', 'OPCO_ID', 'OPCO_ID'),
        (2, 'OPCO_CD', 'OPCO_CD', 'OPCO_CD'),
        (3, 'OPCO_DSC', 'OPCO_DSC', 'OPCO_DSC'),
        (4, 'RPT_OPCO_DSC', 'RPT_OPCO_DSC', 'RPT_OPCO_DSC'),
        (5, 'RPT_OPCO_ABBRV', 'RPT_OPCO_ABBRV', 'RPT_OPCO_ABBRV'),
        (6, 'TBA_ACTIVE_FLG', '', "'Y'"),  # Target-only column
        (7, 'DATA_SRC', '', "'DW.D_OPCO'"),
        (8, 'CREATE_DT', '', 'CURRENT_TIMESTAMP()'),
        (9, 'CREATE_BY', '', "CAST(CURRENT_USER() AS VARCHAR(200))"),
        (10, 'CREATE_PGM', '', "'DW.D_OPCO'"),
        (11, 'UPDATE_DT', '', 'CURRENT_TIMESTAMP()'),
        (12, 'UPDATE_BY', '', "CAST(CURRENT_USER() AS VARCHAR(200))"),
        (13, 'UPDATE_PGM', '', "'DW.D_OPCO'")
    ]
    
    # Add columns to the sheet
    for i, (sno, target, source, logic) in enumerate(columns, 11):
        mapping_sheet.cell(row=i, column=1, value=sno)
        mapping_sheet.cell(row=i, column=2, value=target)
        mapping_sheet.cell(row=i, column=3, value=source)
        mapping_sheet.cell(row=i, column=4, value=logic)
    
    # Add JOIN section
    join_row = 25
    mapping_sheet.cell(row=join_row, column=1, value='JOIN_TABLES')
    
    # Add join headers
    join_headers = ['Join Type', 'Table Type', 'Source Name', 'Table Name', 'Alias', 'Join Condition']
    for col, header in enumerate(join_headers, 1):
        mapping_sheet.cell(row=join_row+1, column=col, value=header)
    
    # Add a sample join
    mapping_sheet.cell(row=join_row+2, column=1, value='LEFT')
    mapping_sheet.cell(row=join_row+2, column=2, value='ref')
    mapping_sheet.cell(row=join_row+2, column=3, value='')
    mapping_sheet.cell(row=join_row+2, column=4, value='LND_CORE.D_ITEM')
    mapping_sheet.cell(row=join_row+2, column=5, value='items')
    mapping_sheet.cell(row=join_row+2, column=6, value='OPCO_ID=items.OPCO_ID')
    
    # Save the workbook
    wb.save(file_path)
    return file_path

def create_test_json_config(file_path, mapping_file):
    """Create a test JSON configuration based on the mapping file"""
    # Create directory if it doesn't exist
    if os.path.dirname(file_path):
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # Load the mapping file
    wb = openpyxl.load_workbook(mapping_file)
    mapping_sheet = wb['Mapping']
    
    # Extract configuration
    target_table = mapping_sheet.cell(row=1, column=2).value
    source_table = mapping_sheet.cell(row=2, column=2).value
    source_type = mapping_sheet.cell(row=3, column=2).value
    source_name = mapping_sheet.cell(row=4, column=2).value
    materialization = mapping_sheet.cell(row=5, column=2).value
    unique_key = mapping_sheet.cell(row=6, column=2).value
    
    # Parse tables
    target_parts = target_table.split('.')
    target_schema = target_parts[0]
    target_table_name = target_parts[1]
    
    source_parts = source_table.split('.')
    source_db = source_parts[0]
    source_schema = source_parts[1]
    source_table_name = source_parts[2]
    
    # Create config structure
    config = {
        'Source': {
            'Type': source_type,
            'Database': source_db,
            'Schema': source_schema,
            'Table Name': source_table_name,
            'Name': source_name
        },
        'Target': {
            'Schema': target_schema,
            'Table Name': target_table_name,
            'materialization': materialization,
            'unique_key': [unique_key] if unique_key else []
        },
        'Columns': []
    }
    
    # Add columns
    for row in range(11, 24):  # Rows 11-23 contain the columns
        target_col = mapping_sheet.cell(row=row, column=2).value
        source_col = mapping_sheet.cell(row=row, column=3).value
        logic = mapping_sheet.cell(row=row, column=4).value
        
        if target_col:
            config['Columns'].append({
                "Target Column": target_col,
                "Source Table": source_col,
                "Logic": logic
            })
    
    # Write the configuration to a file
    with open(file_path, 'w') as f:
        json.dump(config, f, indent=2)
    
    return file_path

def test_model_generation():
    """Test the model generation with our fixes"""
    try:
        # Create test directories
        os.makedirs('models', exist_ok=True)
        
        print("=" * 80)
        print("TESTING MODEL GENERATION WITH FIXES")
        print("=" * 80)
        
        # Create test mapping file with absolute path
        current_dir = os.getcwd()
        mapping_file = os.path.join(current_dir, 'test_mapping.xlsx')
        mapping_file = create_test_mapping_file(mapping_file)
        print(f"Created test mapping file: {mapping_file}")
        
        # Create test JSON config with absolute path
        json_file = os.path.join(current_dir, 'test_config.json')
        json_file = create_test_json_config(json_file, mapping_file)
        print(f"Created test JSON config: {json_file}")
        
        # Generate DBT model
        wb = openpyxl.load_workbook(mapping_file)
        mapping_sheet = wb['Mapping']
        model_file = create_dbt_model_from_json(json_file, mapping_sheet)
        print(f"Generated DBT model: {model_file}")
        
        # Print the contents of the generated model
        print("\n" + "=" * 80)
        print("GENERATED MODEL")
        print("=" * 80)
        
        with open(model_file, 'r') as f:
            content = f.read()
            print(content)
            
        print("\n" + "=" * 80)
        print("VERIFICATION")
        print("=" * 80)
        
        # Check if source alias is added to joining columns
        if "source.OPCO_ID=items.OPCO_ID" in content:
            print("✓ PASSED - Source alias is added to joining columns")
        else:
            print("✗ FAILED - Source alias is not added to joining columns")
            
        # Check if TBA_ACTIVE_FLG is excluded from the MINUS subquery
        if "TBA_ACTIVE_FLG" not in content.split("MINUS")[1].split("FROM")[0]:
            print("✓ PASSED - TBA_ACTIVE_FLG is excluded from the MINUS subquery")
        else:
            print("✗ FAILED - TBA_ACTIVE_FLG is included in the MINUS subquery")
            
        # Check if CREATE_PGM is excluded from merge_update_columns
        if "'CREATE_PGM'" not in content.split("merge_update_columns")[1].split("]")[0]:
            print("✓ PASSED - CREATE_PGM is excluded from merge_update_columns")
        else:
            print("✗ FAILED - CREATE_PGM is included in merge_update_columns")
            
        print("\n" + "=" * 80)
        print("TEST COMPLETED")
        print("=" * 80)
    except Exception as e:
        import traceback
        print(f"ERROR: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    test_model_generation() 