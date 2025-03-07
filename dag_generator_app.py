import datetime
import json
import logging
import os
import re
import subprocess
import sys
import threading
import tkinter as tk
import traceback
import openpyxl
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from ttkthemes import ThemedTk, THEMES

from scripts.dag_generators import (
    create_dataset_dependency_dag,
    create_cron_dag,
    create_sns_dag
)
from scripts.dbt_job_generator import create_dbt_job_file
from scripts.dbt_model_generator import create_dbt_model_from_json
from scripts.insert_sql_generator import insert_sql_generator
from scripts.merge_sql_generator import merge_sql_generator
from scripts.generate_lnd_dbt_model_file import generate_lnd_dbt_model_file


# Configure logging
def setup_logging():
    """Set up logging to file"""
    log_file = os.path.join(os.getcwd(), 'app.log')
    
    # Create a logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    # Create handlers
    file_handler = logging.FileHandler(log_file, mode='a')
    file_handler.setLevel(logging.INFO)
    
    # Create formatters
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    
    # Add handlers to the logger
    logger.addHandler(file_handler)
    
    # Redirect stdout and stderr to the log file
    class LoggerWriter:
        def __init__(self, level):
            self.level = level
            self.buffer = ''

        def write(self, message):
            if message and not message.isspace():
                self.buffer += message
                if self.buffer.endswith('\n'):
                    self.buffer = self.buffer.rstrip()
                    logger.log(self.level, self.buffer)
                    self.buffer = ''

        def flush(self):
            if self.buffer:
                logger.log(self.level, self.buffer)
                self.buffer = ''
    
    # Redirect stdout and stderr to the logger
    sys.stdout = LoggerWriter(logging.INFO)
    sys.stderr = LoggerWriter(logging.ERROR)
    
    return logger

# Initialize logging
logger = setup_logging()

# Lazy imports - only import when needed
def lazy_import():
    global openpyxl, Image, ImageTk, ModelMapper


# Helper function to find similar files when exact match is not found
def find_similar_file(target_path):
    """Find a file with a similar name pattern but different timestamp"""
    # Extract directory and filename pattern
    directory = os.path.dirname(target_path)
    filename = os.path.basename(target_path)
    
    # Extract base name and timestamp
    match = re.match(r'(.+)_(\d{8}_\d{6})\.(.+)', filename)
    if not match:
        return None
    
    base_name, timestamp, extension = match.groups()
    
    # Check if directory exists
    if not os.path.exists(directory):
        return None
    
    # List files in directory
    try:
        files = os.listdir(directory)
        for file in files:
            if file.startswith(base_name) and file.endswith(f".{extension}"):
                return os.path.join(directory, file)
    except Exception:
        pass
    
    return None

# Helper function to ensure file exists or find alternative
def ensure_file_exists(file_path):
    """Check if file exists and try to find alternative if not"""
    if os.path.exists(file_path):
        return file_path
    
    # Try to find a similar file
    similar_file = find_similar_file(file_path)
    if similar_file:
        return similar_file
    
    return file_path  # Return original path if no alternative found

# Helper function to generate a timestamp that's safe for filenames
def generate_safe_timestamp():
    """Generate a timestamp that's safe for filenames and not affected by system date issues"""
    try:
        # Try to get the current date
        now = datetime.datetime.now()
        # Validate the year is reasonable (between 2020 and 2030)
        if now.year < 2020 or now.year > 2030:
            # Use a default year of 2023 if the system date is unreasonable
            now = now.replace(year=2023)
        return now.strftime("%Y%m%d_%H%M%S")
    except Exception:
        # Fallback to a basic timestamp if there's any issue
        return "20230101_000000"

class DAGGeneratorApp:
    def __init__(self, root):
        """Initialize the application"""
        self.primary_keys = None
        self.unique_keys = None
        self.status_bar = None
        self.mapping_tab = None
        self.main_tab = None
        self.notebook = None
        self.ddl_combo = None
        self.mapping_combo = None
        self.animation_duration = 0.5
        self.animation_steps = 120
        self.root = root
        self.root.title("DBT Model Generator")
        self.root.geometry("1366x768")
        
        # Initialize thread pool
        self.executor = ThreadPoolExecutor(max_workers=3)
        
        # Lazy load heavy imports
        self.imports_loaded = False
        
        # # Apply theme immediately
        # style = ttk.Style()
        # try:
        #     # Try to use a themed style if ttkthemes is available
        #     if 'ThemedTk' in globals():
        #         available_themes = THEMES
        #         if 'breeze' in available_themes:
        #             style.theme_use('breeze')
        #         elif 'arc' in available_themes:
        #             style.theme_use('arc')
        #         elif 'clam' in available_themes:
        #             style.theme_use('clam')
        #         else:
        #             # Use default theme
        #             style.theme_use('clam')
        #     else:
        #         # Use default ttk theme
        #         style.theme_use('clam')
        # except Exception as e:
        #     logging.warning(f"Could not set theme: {str(e)}. Using default theme.")
        #     # Use a default theme that's available in standard ttk
        #     try:
        #         style.theme_use('clam')
        #     except:
        #         pass  # If even this fails, just use the default theme
        #
        # Initialize history lists
        self.file_history = []
        self.ddl_file_history = []
        
        # Variables with history
        self.mapping_file_path = tk.StringVar()
        self.json_output_path = tk.StringVar()
        self.model_output_path = tk.StringVar()
        self.dag_output_path = tk.StringVar()
        self.ddl_file_path = tk.StringVar()
        
        # Control variables
        self.generate_dag_var = tk.BooleanVar(value=True)  # Default to True
        self.generate_model_var = tk.BooleanVar(value=True)  # Default to True
        self.generate_dbt_job_var = tk.BooleanVar(value=True)  # Default to True
        self.generate_merge_macro_var = tk.BooleanVar(value=False)  # Default to False
        self.generate_insert_macro_var = tk.BooleanVar(value=False)  # Default to False
        self.generate_lnd_model_var = tk.BooleanVar(value=False)  # Default to False

        # Initialize ModelMapper
        self.model_mapper = None

        # Initialize UI
        self.init_ui()

        # Load history and other resources in background
        self.root.after(100, self.delayed_init)
        

        


    def delayed_init(self):
        """Initialize non-critical components in background"""
        thread = threading.Thread(target=self.load_background_resources)
        thread.daemon = True
        thread.start()

    def load_background_resources(self):
        """Load resources that aren't needed immediately"""
        try:
            # Load imports
            lazy_import()
            self.imports_loaded = True
            
            # Initialize ModelMapper after imports are loaded
            from scripts.model_mapper import ModelMapper
            self.model_mapper = ModelMapper()
            
            # Load history
            self.load_history()
            
            # Update UI with loaded resources
            self.root.after(0, self.update_ui_with_resources)
        except Exception as e:
            logging.error(f"Background loading error: {str(e)}")

    @lru_cache(maxsize=32)
    def load_history(self):
        """Load file history from a JSON file with caching"""
        try:
            history_file = os.path.join(os.getcwd(), 'file_history.json')
            if os.path.exists(history_file):
                with open(history_file, 'r') as f:
                    history = json.load(f)
                    
                    # Validate file paths in history
                    self.file_history = [path for path in history.get('mapping_files', []) 
                                        if os.path.exists(path) or find_similar_file(path)]
                    
                    self.ddl_file_history = [path for path in history.get('ddl_files', []) 
                                            if os.path.exists(path) or find_similar_file(path)]
                    return True
        except:
            self.file_history = []
            self.ddl_file_history = []
        return False

    def init_ui(self):
        """Initialize the basic UI structure"""
        # Create main container
        main_container = ttk.Frame(self.root, padding="20")
        main_container.pack(expand=True, fill='both')

        # Create header
        ttk.Label(
            main_container,
            text="DAG Generator",
            style='Header.TLabel'
        ).pack(anchor='w', pady=(0, 5))

        ttk.Label(
            main_container,
            text="Generate DBT models, jobs and Airflow DAGs",
            style='Subheader.TLabel'
        ).pack(anchor='w', pady=(0, 20))

        # Create notebook and tabs
        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(expand=True, fill='both')

        # Create basic tab frames
        self.main_tab = ttk.Frame(self.notebook, padding=20)
        self.mapping_tab = ttk.Frame(self.notebook, padding=20)

        self.notebook.add(self.main_tab, text='DAG Generator')
        self.notebook.add(self.mapping_tab, text='Mapping Tools')

        # Initialize basic tab content
        self.init_main_tab()
        self.init_mapping_tab()
        
        # Create status bar
        self.status_bar = ttk.Frame(self.root)
        self.status_bar.pack(fill='x', side='bottom')
        
        self.status_label = ttk.Label(self.status_bar, text="Ready", anchor='w')
        self.status_label.pack(side='left', padx=10)
        
        self.progress = ttk.Progressbar(self.status_bar, mode='indeterminate', length=100)
        self.progress.pack(side='right', padx=10, pady=5)
        
        # Hide progress bar initially
        self.progress.pack_forget()
        
        # Apply custom styles
        self.apply_custom_styles()

    def apply_custom_styles(self):
        """Apply custom styles to the UI elements"""
        style = ttk.Style()
        
        # Header style
        style.configure('Header.TLabel', font=('Segoe UI', 16, 'bold'))
        
        # Subheader style
        style.configure('Subheader.TLabel', font=('Segoe UI', 10))
        
        # Accent button style
        style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))
        
        # Normal button style
        style.configure('TButton', font=('Segoe UI', 10))
        
        # Label style
        style.configure('TLabel', font=('Segoe UI', 10))
        
        # Entry style
        style.configure('TEntry', font=('Segoe UI', 10))
        
        # Combobox style
        style.configure('TCombobox', font=('Segoe UI', 10))

    def init_main_tab(self):
        """Initialize DAG Generator tab"""
        # Create sections
        generate_frame = ttk.LabelFrame(
            self.main_tab, 
            text="Generate from Mapping",
            padding=15
        )
        generate_frame.pack(fill='x', pady=(0, 20))

        # File selection frame with Combobox
        file_frame = ttk.Frame(generate_frame)
        file_frame.pack(fill='x', pady=(0, 15))
        
        ttk.Label(
            file_frame, 
            text="Upload Excel File:",
            width=15  # Fixed width for label
        ).pack(side='left', padx=(0, 10))
        
        # Container frame for combobox and button
        input_frame = ttk.Frame(file_frame)
        input_frame.pack(side='left', fill='x', expand=True)
        
        # Combobox with adjusted width
        self.file_combo = ttk.Combobox(
            input_frame,
            textvariable=self.mapping_file_path,
            values=self.file_history,
            width=50  # Adjusted width
        )
        self.file_combo.pack(side='left', fill='x', expand=True, padx=(0, 10))
        
        # Browse button with fixed width
        ttk.Button(
            input_frame, 
            text="Browse",
            command=self.browse_file,
            width=10  # Fixed width for button
        ).pack(side='right')

        self.create_enhanced_tooltip(self.file_combo, "Select an Excel mapping file")

        # Bind events
        self.file_combo.bind('<<ComboboxSelected>>', self.on_combo_select)

        # Options frame
        options_frame = ttk.Frame(generate_frame)
        options_frame.pack(fill='x', pady=(5, 10))
        
        # Generate DAG checkbox
        ttk.Checkbutton(
            options_frame,
            text="Generate DAG file",
            variable=self.generate_dag_var
        ).pack(side='left', padx=(0, 10))

        # Generate Model checkbox
        ttk.Checkbutton(
            options_frame,
            text="Generate Model file",
            variable=self.generate_model_var
        ).pack(side='left', padx=(0, 10))

        # Generate dbt job checkbox
        ttk.Checkbutton(
            options_frame,
            text="Generate dbt job file",
            variable=self.generate_dbt_job_var
        ).pack(side='left', padx=(0, 10))

        # Generate insert macro checkbox
        ttk.Checkbutton(
            options_frame,
            text="Generate insert macro file",
            variable=self.generate_insert_macro_var
        ).pack(side='left', padx=(0, 10))

        # Generate merge_macro checkbox
        ttk.Checkbutton(
            options_frame,
            text="Generate merge macro file",
            variable=self.generate_merge_macro_var
        ).pack(side='left', padx=(0, 10))

        # Generate lnd_model checkbox
        ttk.Checkbutton(
            options_frame,
            text="Generate LND Model file",
            variable=self.generate_lnd_model_var
        ).pack(side='left', padx=(0, 10))

        # Generate button
        self.generate_button = ttk.Button(
            generate_frame,
            text="Generate Files",
            command=self.start_generation,
            style='Accent.TButton'
        )
        self.generate_button.pack(fill='x', pady=(10, 0))

        # DAG Directory section
        dir_frame = ttk.LabelFrame(
            self.main_tab, 
            text="DAG Directory",
            padding=15
        )
        dir_frame.pack(fill='x', pady=(0, 20))

        ttk.Button(
            dir_frame,
            text="Open DAG Directory",
            command=self.open_dag_directory
        ).pack(fill='x')
        
        # Help section
        help_frame = ttk.LabelFrame(
            self.main_tab,
            text="Help",
            padding=15
        )
        help_frame.pack(fill='x')
        
        help_text = (
            "1. Select a mapping Excel file using the Browse button or dropdown\n"
            "2. Click 'Generate Files' to create DBT model, job, and DAG files\n"
            "3. Use 'Open DAG Directory' to view the generated DAG files\n\n"
            "For more help, see the README.md file."
        )
        
        ttk.Label(
            help_frame,
            text=help_text,
            wraplength=700,
            justify='left'
        ).pack(fill='x')

    def init_mapping_tab(self):
        """Initialize Mapping Tools tab"""
        # DDL Section
        ddl_frame = ttk.LabelFrame(
            self.mapping_tab, 
            text="Generate Mapping from DDL",
            padding=15
        )
        ddl_frame.pack(fill='x', pady=(0, 20))

        # DDL File selection
        ddl_file_frame = ttk.Frame(ddl_frame)
        ddl_file_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(
            ddl_file_frame, 
            text="Select DDL File:",
            width=15
        ).pack(side='left', padx=(0, 10))
        
        # Container for combobox and button
        input_frame = ttk.Frame(ddl_file_frame)
        input_frame.pack(side='left', fill='x', expand=True)
        
        self.ddl_combo = ttk.Combobox(
            input_frame,
            textvariable=self.ddl_file_path,
            values=self.ddl_file_history,
            width=50
        )
        self.ddl_combo.pack(side='left', fill='x', expand=True, padx=(0, 10))
        
        ttk.Button(
            input_frame, 
            text="Browse",
            command=self.browse_ddl_file,
            width=10
        ).pack(side='right')

        self.create_enhanced_tooltip(self.ddl_combo, "Select a SQL DDL file to generate a mapping template")

        # Generate Mapping Button
        self.generate_mapping_button = ttk.Button(
            ddl_frame,
            text="Generate Mapping",
            command=self.generate_mapping_from_ddl,
            style='Accent.TButton'
        )
        self.generate_mapping_button.pack(fill='x', pady=(10, 0))
        
        # Help text for DDL section
        ttk.Label(
            ddl_frame,
            text="This will create an Excel template from a DDL file. You'll need to fill in the SOURCE_TABLE field manually.",
            wraplength=700,
            justify='left'
        ).pack(fill='x', pady=(10, 0))

        # Model Mapping Section
        mapping_frame = ttk.LabelFrame(
            self.mapping_tab, 
            text="Fill Model Mapping",
            padding=15
        )
        mapping_frame.pack(fill='x')

        # Mapping file selection with Combobox
        mapping_file_frame = ttk.Frame(mapping_frame)
        mapping_file_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(
            mapping_file_frame, 
            text="Select Mapping File:"
        ).pack(side='left')
        
        # Replace Entry with Combobox
        self.mapping_combo = ttk.Combobox(
            mapping_file_frame,
            textvariable=self.mapping_file_path,
            values=self.file_history,
            width=50
        )
        self.mapping_combo.pack(side='left', expand=True, fill='x', padx=10)
        
        ttk.Button(
            mapping_file_frame, 
            text="Browse",
            command=self.browse_file
        ).pack(side='right')
        
        self.create_enhanced_tooltip(self.mapping_combo, "Select a mapping file that has SOURCE_TABLE filled in")

        # Fill Model Mapping Button
        self.fill_mapping_button = ttk.Button(
            mapping_frame,
            text="Fill Model Mapping",
            command=self.generate_model_mapping,
            style='Accent.TButton'
        )
        self.fill_mapping_button.pack(fill='x', pady=(10, 0))
        
        # Help text for Model Mapping section
        ttk.Label(
            mapping_frame,
            text="This will connect to Snowflake to fill in column details. Make sure SOURCE_TABLE is filled in the mapping file.",
            wraplength=700,
            justify='left'
        ).pack(fill='x', pady=(10, 0))

    def start_generation(self):
        """Start the generation process with progress indicator"""
        self.set_status("Generating files...")
        self.show_progress()
        self.generate_button.configure(state='disabled')
        
        # Run in a separate thread to keep UI responsive
        threading.Thread(target=self.run_generation).start()
    
    def run_generation(self):
        """Run the generation process in a separate thread"""
        try:
            self.generate_from_mapping()
        finally:
            # Re-enable button and hide progress
            self.root.after(0, self.finish_generation)
    
    def finish_generation(self):
        """Clean up after generation process completes"""
        self.generate_button.configure(state='normal')
        self.hide_progress()
        self.set_status("Ready")

    def set_status(self, message):
        """Update status bar message"""
        self.status_label.config(text=message)
        
    def show_progress(self):
        """Show progress bar"""
        self.progress.pack(side='right', padx=10, pady=5)
        self.progress.start(10)
        
    def hide_progress(self):
        """Hide progress bar"""
        self.progress.stop()
        self.progress.pack_forget()

    def create_enhanced_tooltip(self, widget, tooltip_text=None):
        """Create enhanced tooltip for showing full path and custom text"""
        tooltip = tk.Label(
            self.root,
            text='',
            background='#ffffe0',
            relief='solid',
            borderwidth=1,
            font=('Segoe UI', 9),
            padx=5,
            pady=3
        )
        
        def show_tooltip(event):
            # Get the full path from the actual storage, not the display value
            full_path = self.mapping_file_path.get()
            if not full_path and hasattr(widget, 'get'):
                full_path = widget.get()
            
            # Use custom tooltip text if provided
            if tooltip_text and not full_path:
                tooltip.config(text=tooltip_text)
            elif full_path:
                tooltip.config(text=full_path)
            else:
                return  # No tooltip to show
                
            tooltip.lift()
            
            # Position tooltip below the widget
            x = widget.winfo_rootx()
            y = widget.winfo_rooty() + widget.winfo_height() + 5
            
            # Ensure tooltip doesn't go off screen
            tooltip_width = len(tooltip.cget('text')) * 7  # Approximate width calculation
            screen_width = widget.winfo_screenwidth()
            if x + tooltip_width > screen_width:
                x = screen_width - tooltip_width - 10
            
            tooltip.place(x=x, y=y)
        
        def hide_tooltip(event):
            tooltip.place_forget()
        
        widget.bind('<Enter>', show_tooltip)
        widget.bind('<Leave>', hide_tooltip)
        widget._tooltip = tooltip  # Store reference to prevent garbage collection

    def browse_file(self):
        """Browse for Excel file"""
        # Ensure mappings directory exists
        initial_dir = os.path.join(os.getcwd(), "mappings")
        os.makedirs(initial_dir, exist_ok=True)
        
        file_path = filedialog.askopenfilename(
            title="Select Mapping File",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=initial_dir
        )
        if file_path:
            self.mapping_file_path.set(file_path)
            self.add_to_history(file_path, self.file_history)
            self.file_combo['values'] = self.file_history
            if hasattr(self, 'mapping_combo'):
                self.mapping_combo['values'] = self.file_history

    def generate_from_mapping(self):
        """Generate DBT model and DAG from mapping file"""
        try:
            # Validate inputs
            if not self.mapping_file_path.get():
                raise ValueError("Please select a mapping file")
                
            if not os.path.exists(self.mapping_file_path.get()):
                raise ValueError(f"Mapping file not found: {self.mapping_file_path.get()}")
                
            # Get output paths
            json_output_path = self.json_output_path.get() or os.path.join('data', 'model_config.json')
            model_output_path = self.model_output_path.get() or 'models'
            dag_output_path = self.dag_output_path.get() or 'dags'
            
            # Ensure directories exist
            os.makedirs(os.path.dirname(json_output_path), exist_ok=True)
            os.makedirs(model_output_path, exist_ok=True)
            os.makedirs(dag_output_path, exist_ok=True)
            
            # Load mapping file
            workbook = load_workbook(self.mapping_file_path.get())
            mapping_sheet = workbook['Mapping']
            
            # Extract table information
            target_table = None
            source_table = None
            source_type = None
            source_name = None
            materialization = None
            
            for row in range(1, mapping_sheet.max_row + 1):
                cell_value = mapping_sheet.cell(row=row, column=1).value
                if cell_value == 'TARGET_TABLE':
                    target_table = mapping_sheet.cell(row=row, column=2).value
                elif cell_value == 'SOURCE_TABLE':
                    source_table = mapping_sheet.cell(row=row, column=2).value
                elif cell_value == 'SOURCE_TYPE':
                    source_type = mapping_sheet.cell(row=row, column=2).value
                elif cell_value == 'SOURCE_NAME':
                    source_name = mapping_sheet.cell(row=row, column=2).value
                elif cell_value == 'MATERIALIZATION':
                    materialization = mapping_sheet.cell(row=row, column=2).value
                
                if target_table and source_table and source_type and source_name and materialization:
                    break
            
            if not target_table:
                raise ValueError("Target table not found in mapping file")
            if not source_table:
                raise ValueError("Source table not found in mapping file")
            if not source_type:
                source_type = 'source'  # Default to source
            if not source_name:
                # Extract schema from source table as default source name
                parts = source_table.split('.')
                if len(parts) > 1:
                    source_name = parts[1]
                else:
                    source_name = 'default_source'
            if not materialization:
                materialization = 'incremental'  # Default to incremental
                
            # Parse table names
            source_parts = source_table.split('.')
            if len(source_parts) < 3:
                raise ValueError("Source table should be in format: DATABASE.SCHEMA.TABLE")
                
            source_db = source_parts[0]
            source_schema = source_parts[1]
            source_table_name = source_parts[2]
            
            target_parts = target_table.split('.')
            if len(target_parts) < 2:
                raise ValueError("Target table should be in format: SCHEMA.TABLE")
                
            target_schema = target_parts[0]
            target_table_name = target_parts[1]
            
            # Extract column mappings
            column_mappings = []
            header_row = None
            
            # Find the header row
            for row in range(1, mapping_sheet.max_row + 1):
                if mapping_sheet.cell(row=row, column=1).value == 'S.NO':
                    header_row = row
                    break
                    
            if not header_row:
                raise ValueError("Column mapping header not found in mapping file")
                
            # Extract column mappings
            for row in range(header_row + 1, mapping_sheet.max_row + 1):
                target_column = mapping_sheet.cell(row=row, column=2).value
                source_table_col = mapping_sheet.cell(row=row, column=3).value
                logic = mapping_sheet.cell(row=row, column=4).value
                
                if not target_column:
                    # Stop when we reach an empty row or the JOIN_TABLES section
                    next_row_value = mapping_sheet.cell(row=row+1, column=1).value
                    if next_row_value and next_row_value == 'JOIN_TABLES':
                        break
                    continue
                    
                column_mappings.append({
                    'Target Column': target_column,
                    'Source Table': source_table_col,
                    'Logic': logic or target_column  # Use target column as default logic
                })
            
            # Create model configuration
            model_config = {
                'Source': {
                    'Type': source_type,
                    'Database': source_db,
                    'Schema': source_schema,
                    'Table Name': source_table_name,
                    'Name': source_name
                },
                'Target': {
                    'Schema': target_schema,
                    'Table Name': target_table_name,
                    'materialization': materialization
                },
                'Columns': column_mappings
            }
            
            # Check if we need to add unique keys for incremental models
            if materialization == 'incremental':
                # Try to get unique keys from the DDL file if we have it
                unique_keys = []
                primary_keys = []
                ddl_unique_keys_found = False
                
                # First check if we have a DDL file path directly specified
                ddl_path = None
                if hasattr(self, 'ddl_file_path') and self.ddl_file_path.get():
                    ddl_path = self.ddl_file_path.get()
                    logging.info(f"Using explicitly specified DDL file: {ddl_path}")
                else:
                    # Try to find a matching DDL file in history
                    logging.info("Looking for matching DDL file in history...")
                    # Extract table name from target table
                    target_table_name = target_table_name.upper()
                    
                    # Look through DDL file history for a matching file
                    for history_ddl_path in self.ddl_file_history:
                        if os.path.exists(history_ddl_path):
                            # Check if the DDL file contains the target table name
                            try:
                                with open(history_ddl_path, 'r') as f:
                                    ddl_content = f.read()
                                    # Simple check if the table name appears in the DDL
                                    if f"CREATE TABLE {target_schema}.{target_table_name}" in ddl_content.upper() or \
                                       f"CREATE TABLE \"{target_schema}\".\"{target_table_name}\"" in ddl_content.upper():
                                        ddl_path = history_ddl_path
                                        logging.info(f"Found matching DDL file in history: {ddl_path}")
                                        break
                            except Exception as e:
                                logging.error(f"Error checking DDL file {history_ddl_path}: {str(e)}")
                
                if ddl_path and os.path.exists(ddl_path):
                    logging.info(f"DDL file exists: {ddl_path}")
                    try:
                        # Log first few lines of the DDL file for debugging
                        with open(ddl_path, 'r') as f:
                            ddl_content = f.read(500)  # Read first 500 chars
                            logging.debug(f"DDL file content (first 500 chars):\n{ddl_content}")
                        
                        _, unique_keys = self.parse_ddl_file(ddl_path)
                        if hasattr(self, 'primary_keys'):
                            primary_keys = self.primary_keys
                        
                        # If we found unique keys in the DDL, mark that we should use these directly
                        if unique_keys:
                            ddl_unique_keys_found = True
                            logging.info(f"Using unique keys from DDL: {unique_keys}")
                        else:
                            logging.info("No unique keys found in DDL")
                            
                        if primary_keys:
                            logging.info(f"Found primary keys in DDL: {primary_keys}")
                    except Exception as e:
                        logging.error(f"Error extracting keys from DDL: {str(e)}")
                else:
                    logging.info("No matching DDL file found in history")
                
                # If we have unique keys stored from previous parsing, use those
                if not unique_keys and hasattr(self, 'unique_keys') and self.unique_keys:
                    unique_keys = self.unique_keys
                    if unique_keys:
                        ddl_unique_keys_found = True
                        logging.info(f"Using stored unique keys from DDL: {unique_keys}")
                
                # Check if unique keys are specified in the mapping sheet
                if not ddl_unique_keys_found:
                    try:
                        # Look for UNIQUE_KEY in the mapping sheet
                        for row in range(1, 10):  # Check first few rows
                            if mapping_sheet.cell(row=row, column=1).value == 'UNIQUE_KEY':
                                uk_value = mapping_sheet.cell(row=row, column=2).value
                                if uk_value:
                                    # Split by comma if multiple keys
                                    if ',' in uk_value:
                                        unique_keys = [key.strip() for key in uk_value.split(',')]
                                    else:
                                        unique_keys = [uk_value.strip()]
                                    logging.info(f"Found unique keys in mapping sheet: {unique_keys}")
                                break
                    except Exception as e:
                        logging.error(f"Error checking for unique keys in mapping sheet: {str(e)}")
                
                # Only if we didn't find unique keys in DDL or mapping sheet, try pattern-based detection
                if not unique_keys and not ddl_unique_keys_found:
                    # First, check for columns with _CD suffix (common for unique business keys)
                    cd_columns = []
                    for column in column_mappings:
                        target_col = column['Target Column']
                        if target_col.endswith('_CD'):
                            cd_columns.append(target_col)
                    
                    if cd_columns:
                        unique_keys = cd_columns
                        logging.info(f"Using _CD columns as unique keys: {unique_keys}")
                    else:
                        # Look for ID columns or primary key candidates
                        id_columns = []
                        for column in column_mappings:
                            target_col = column['Target Column']
                            # Check for common ID column patterns
                            if (target_col.endswith('_ID') or 
                                target_col.endswith('_KEY') or 
                                target_col == 'ID' or 
                                target_col == 'KEY'):
                                id_columns.append(target_col)
                        
                        if id_columns:
                            unique_keys = id_columns
                            logging.info(f"Using ID columns as unique keys: {unique_keys}")
                        elif primary_keys:
                            # Use primary keys if no other unique keys found
                            unique_keys = primary_keys
                            logging.info(f"Using primary keys as unique keys: {unique_keys}")
                        else:
                            # If no ID columns found, use the first column as a fallback
                            if column_mappings:
                                unique_keys = [column_mappings[0]['Target Column']]
                                logging.info(f"Using first column as unique key: {unique_keys}")
                
                # Add unique keys to the model config if we have any
                if unique_keys:
                    logging.info(f"Adding unique keys to model config: {unique_keys}")
                    model_config['Target']['unique_key'] = unique_keys
            
            # Save model configuration to JSON
            with open(json_output_path, 'w') as f:
                json.dump(model_config, f, indent=2)
                
            # Generate DBT model
            model_file_path = None
            if self.generate_model_var.get():
                model_file_path = create_dbt_model_from_json(json_output_path, mapping_sheet, self.ddl_file_path.get())
            
            # Generate DAG file if requested
            dag_file_path = None
            if self.generate_dag_var.get():
                dag_file_path = self.generate_dag_file(json_output_path, dag_output_path)

            # Generate merge_macro file if requested
            merge_macro_file_path = None
            merge_dbt_job_additon_flg=False
            if self.generate_merge_macro_var.get():
                merge_dbt_job_additon_flg,merge_macro_file_path = merge_sql_generator(json_output_path,mapping_sheet, merge_macro_file_path)

            # Generate insert_macro file if requested
            insert_macro_file_path = None
            insert_dbt_job_additon_flg = False
            if self.generate_insert_macro_var.get():
                insert_dbt_job_additon_flg,insert_macro_file_path = insert_sql_generator(json_output_path,mapping_sheet, insert_macro_file_path)

            # Generate lnd_model file if requested
            lnd_model_file_path = None
            if self.generate_lnd_model_var.get():
                print(json_output_path)
                lnd_model_file_path = generate_lnd_dbt_model_file(json_output_path, self.mapping_file_path.get())

            # Generate DBT job file
            job_output_path = None
            if self.generate_dbt_job_var.get():
                job_output_path = 'jobs'
                job_output_path = create_dbt_job_file(json_output_path, job_output_path, merge_dbt_job_additon_flg, merge_macro_file_path, insert_dbt_job_additon_flg, insert_macro_file_path)

            # Prepare success message
            success_message = "Files generated successfully!\n\n"
            if dag_file_path:
                success_message += f"DAG file: {dag_file_path}\n"
            if model_file_path:
                success_message += f"Model file: {model_file_path}\n"
            if job_output_path:
                success_message += f"DBT job file: {job_output_path}\n"
            if merge_macro_file_path:
                success_message += f"Merge Macro file: {merge_macro_file_path}\n"
            if insert_macro_file_path:
                success_message += f"Insert Macro file: {insert_macro_file_path}\n"
            if lnd_model_file_path:
                success_message += f"LND Model file: {lnd_model_file_path}\n"
                
            # Show success message
            messagebox.showinfo("Success", success_message)
            
            # Add to history
            self.add_to_history(self.mapping_file_path.get(), self.file_history)
            self.save_history()
            
            # Update status
            self.set_status("Files generated successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.set_status(f"Error: {str(e)}")

    def generate_dag_file(self, json_output_path, dag_output_path):
        """Generate the appropriate DAG file based on configuration"""
        try:
            with open(json_output_path, 'r') as json_file:
                config = json.load(json_file)
                
            # Check if DAG key exists, if not, use default values
            if 'DAG' not in config:
                # Try to read DAG configuration from the Config sheet in the mapping file
                dag_config = self.get_dag_config_from_mapping()
                logging.info(f"DAG config from mapping: {dag_config}")
                
                # If we couldn't get DAG config from mapping, use default
                if not dag_config:
                    dag_config = {
                        "Type": "dataset_dependency",  # Default to dataset_dependency
                        "Schedule": "0 */4 * * *"      # Default schedule
                    }
                    logging.info(f"Using default DAG config: {dag_config}")
                
                # Add DAG config to the model config
                config['DAG'] = dag_config
                
                # Save the updated config back to the file
                with open(json_output_path, 'w') as f:
                    json.dump(config, f, indent=2)
            
            # Get DAG type from config (with default)
            dag_type = config['DAG'].get("Type", "dataset_dependency").lower()
            logging.info(f"DAG type: {dag_type}")
            logging.info(f"DAG config: {config['DAG']}")
            
            # Map generator functions to the normalized DAG types
            dag_generators = {
                "dataset_dependency": create_dataset_dependency_dag,
                "dataset dependency": create_dataset_dependency_dag,
                "cron": create_cron_dag,
                "sns": create_sns_dag
            }
            
            # Get the appropriate generator function
            generator = dag_generators.get(dag_type)
            if not generator:
                raise ValueError(f"Unsupported DAG Type: {dag_type}")
            
            # Create a proper output file path
            # Extract table name from config
            schema_name = config['Target']['Schema']
            table_name = config['Target']['Table Name']
            
            # Create DAG filename
            dag_filename = f"{schema_name}_{table_name}_dag.py"
            dag_file_path = os.path.join(dag_output_path, dag_filename)
            
            # Create the DAG using the selected generator
            logging.info(f"Generating {dag_type} DAG with config: {config['DAG']}")
            generator(json_output_path, dag_file_path)
            logging.info(f"DAG file generated: {dag_file_path}")
            
            # Return the path to the generated DAG file
            return dag_file_path
            
        except Exception as e:
            error_msg = f"An error occurred while generating the {dag_type.upper()} DAG: {str(e)}"
            messagebox.showerror("DAG Generation Error", error_msg)
            self.set_status(error_msg)
            logging.error(error_msg)
            return None

    def get_dag_config_from_mapping(self):
        """Extract DAG configuration from the mapping file's Config sheet"""
        try:
            if not self.mapping_file_path.get() or not os.path.exists(self.mapping_file_path.get()):
                logging.warning("Mapping file path is empty or file does not exist")
                return None
                
            workbook = load_workbook(self.mapping_file_path.get())
            if 'Config' not in workbook.sheetnames:
                logging.warning("Config sheet not found in mapping file")
                return None
                
            config_sheet = workbook['Config']
            dag_config = {}
            
            # Look for DAG Type and Schedule Interval
            for row in range(3, 10):  # Check first few rows
                param = config_sheet.cell(row=row, column=1).value
                if not param:
                    continue
                    
                if param == 'DAG Type':
                    dag_type = config_sheet.cell(row=row, column=2).value
                    if dag_type:
                        # Normalize DAG type (case-insensitive)
                        dag_type = dag_type.strip().upper()
                        # Map to expected values
                        type_mapping = {
                            'DATASET DEPENDENCY': 'dataset_dependency',
                            'DATASET': 'dataset_dependency',
                            'CRON': 'cron',
                            'SNS': 'sns'
                        }
                        normalized_type = type_mapping.get(dag_type, 'dataset_dependency')
                        dag_config['Type'] = normalized_type
                        logging.info(f"Found DAG Type in mapping: {dag_type} (normalized to {normalized_type})")
                elif param == 'Schedule Interval':
                    schedule = config_sheet.cell(row=row, column=2).value
                    if schedule:
                        dag_config['Schedule'] = schedule.strip()
                        logging.info(f"Found Schedule Interval in mapping: {schedule}")
            
            # If we found a DAG type but no schedule, set a default schedule based on the type
            if 'Type' in dag_config and 'Schedule' not in dag_config:
                if dag_config['Type'] == 'cron':
                    dag_config['Schedule'] = '0 */4 * * *'  # Default cron schedule
                    logging.info(f"Setting default schedule for cron DAG: {dag_config['Schedule']}")
            
            # Look for dependencies if this is a dataset_dependency DAG
            if dag_config.get('Type') == 'dataset_dependency':
                # Find the Dependencies section
                dependency_row = None
                for row in range(3, config_sheet.max_row):
                    if config_sheet.cell(row=row, column=1).value == 'Dependencies':
                        dependency_row = row + 2  # Skip the header row
                        break
                
                if dependency_row:
                    dependency_schemas = []
                    dependency_objects = []
                    
                    # Read up to 5 rows of dependencies
                    for row in range(dependency_row, dependency_row + 5):
                        schema = config_sheet.cell(row=row, column=2).value
                        table = config_sheet.cell(row=row, column=3).value
                        
                        if schema and table:
                            dependency_schemas.append(schema)
                            dependency_objects.append(table)
                            logging.info(f"Found dependency: {schema}.{table}")
                    
                    if dependency_schemas and dependency_objects:
                        dag_config['Dependency Schema'] = dependency_schemas
                        dag_config['Dependency Object'] = dependency_objects
                        logging.info(f"Added {len(dependency_schemas)} dependencies to DAG config")
                    else:
                        logging.warning("No dependencies found for dataset_dependency DAG")
            
            return dag_config
        except Exception as e:
            logging.error(f"Error extracting DAG config from mapping: {str(e)}")
            return None

    def open_dag_directory(self):
        dag_directory = os.path.abspath("dags")
        if os.path.exists(dag_directory):
            if os.name == 'nt':  # For Windows
                os.startfile(dag_directory)
            elif os.name == 'posix':  # For macOS and Linux
                subprocess.call(['open', dag_directory])
            else:
                messagebox.showerror("Error", "Unsupported operating system.")
        else:
            messagebox.showerror("Error", "DAG directory does not exist.")

    def browse_ddl_file(self):
        """Browse for DDL file"""
        file_path = filedialog.askopenfilename(
            title="Select DDL File",
            filetypes=[("SQL files", "*.sql"), ("All files", "*.*")]
        )
        if file_path:
            self.ddl_file_path.set(file_path)
            self.add_to_history(file_path, self.ddl_file_history)
            self.ddl_combo['values'] = self.ddl_file_history
            
            # Create mappings directory if it doesn't exist
            # Use a more reliable path - create mappings in the current workspace
            mappings_dir = os.path.join(os.getcwd(), "mappings")
            os.makedirs(mappings_dir, exist_ok=True)
            
            # Generate default Excel file path
            ddl_name = os.path.splitext(os.path.basename(file_path))[0]
            default_excel_path = os.path.join(mappings_dir, f"{ddl_name}_mapping.xlsx")
            
            if os.path.exists(default_excel_path):
                # Ask user what to do with existing file
                response = messagebox.askyesnocancel(
                    "File Exists",
                    f"Mapping file already exists:\n{default_excel_path}\n\n"
                    "Yes: Replace existing file\n"
                    "No: Create new file with timestamp\n"
                    "Cancel: Keep existing file"
                )
                
                if response is None:  # Cancel
                    self.mapping_file_path.set(default_excel_path)
                elif response:  # Yes - Replace
                    self.mapping_file_path.set(default_excel_path)
                else:  # No - Create new
                    timestamp = generate_safe_timestamp()
                    new_excel_path = os.path.join(mappings_dir, f"{ddl_name}_mapping_{timestamp}.xlsx")
                    self.mapping_file_path.set(new_excel_path)
            else:
                self.mapping_file_path.set(default_excel_path)

    def generate_mapping_from_ddl(self):
        """Generate mapping template from DDL without filling model details"""
        if not self.ddl_file_path.get():
            messagebox.showerror("Error", "Please select a DDL file")
            return
            
        if not os.path.exists(self.ddl_file_path.get()):
            messagebox.showerror("Error", "DDL file does not exist")
            return

        # Show progress
        self.set_status("Generating mapping template...")
        self.show_progress()
        self.generate_mapping_button.configure(state='normal')
        
        # Run in a separate thread
        threading.Thread(target=self.run_mapping_generation).start()
    
    def run_mapping_generation(self):
        """Run the mapping generation in a separate thread"""
        try:
            # Parse DDL file to extract columns
            columns, unique_keys = self.parse_ddl_file(self.ddl_file_path.get())
            
            # Store unique keys for later use
            self.unique_keys = unique_keys
            
            # Update the mapping sheet with columns
            self.update_mapping_sheet((columns, unique_keys))
            
            # Update status
            self.set_status("Mapping file created successfully!")
            self.hide_progress()
            
        except Exception as e:
            self.set_status(f"Error: {str(e)}")
            self.hide_progress()
            messagebox.showerror("Error", str(e))

    def finish_mapping_generation(self):
        """Clean up after mapping generation completes"""
        self.generate_mapping_button.configure(state='normal')
        self.hide_progress()
        self.set_status("Ready")

    def generate_model_mapping(self):
        """Generate model mapping using Snowflake connection"""
        if not self.mapping_file_path.get():
            messagebox.showerror("Error", "Please select Excel file")
            return
            
        # Show progress
        self.set_status("Connecting to Snowflake and updating mapping...")
        self.show_progress()
        self.fill_mapping_button.configure(state='disabled')
        
        # Run in a separate thread
        threading.Thread(target=self.run_model_mapping).start()
    
    def run_model_mapping(self):
        """Run the model mapping in a separate thread"""
        try:
            # Ensure the file path exists
            file_path = self.mapping_file_path.get()
            
            # Use the helper function to find an existing file if the exact path doesn't exist
            file_path = ensure_file_exists(file_path)
            
            directory = os.path.dirname(file_path)
            
            # Create directory if it doesn't exist
            if not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
                
            # Check if the file exists
            if not os.path.exists(file_path):
                self.root.after(0, lambda: messagebox.showerror("Error", f"Mapping file not found: {file_path}"))
                return
            
            # Validate the Excel file has the required sheets and fields
            try:
                workbook = load_workbook(file_path)
                if 'Mapping' not in workbook.sheetnames or 'Config' not in workbook.sheetnames:
                    self.root.after(0, lambda: messagebox.showerror("Error", "Required sheets (Mapping and Config) not found in Excel"))
                    return
                
                mapping_sheet = workbook['Mapping']
                
                # Check for SOURCE_TABLE
                source_table = None
                for row in range(1, 6):
                    if mapping_sheet.cell(row=row, column=1).value == 'SOURCE_TABLE':
                        source_table = mapping_sheet.cell(row=row, column=2).value
                        break
                
                if not source_table:
                    self.root.after(0, lambda: messagebox.showerror("Error", "SOURCE_TABLE is missing in the mapping sheet. Please fill it in before generating the model mapping."))
                    return
                
                # Validate SOURCE_TABLE format
                try:
                    parts = source_table.split('.')
                    if len(parts) != 3:
                        self.root.after(0, lambda: messagebox.showerror("Error", "SOURCE_TABLE should be in format: DATABASE.SCHEMA.TABLE"))
                        return
                except:
                    self.root.after(0, lambda: messagebox.showerror("Error", "Invalid SOURCE_TABLE format"))
                    return
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Error validating Excel file: {str(e)}"))
                return
            
            # Call the model mapper to fill in the model details
            try:
                mapped_count, total_count = self.model_mapper.generate_model_mapping(file_path)
                
                self.root.after(0, lambda: messagebox.showinfo(
                    "Success", 
                    f"Model mapping updated successfully!\n"
                    f"Mapped {mapped_count} out of {total_count} columns."
                ))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Error generating model mapping: {str(e)}"))
                logging.error(f"Error in model_mapper.generate_model_mapping: {str(e)}")
                import traceback
                logging.error(traceback.format_exc())
                
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred: {str(e)}"))
            logging.error(f"Error in run_model_mapping: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
        finally:
            # Re-enable button and hide progress
            self.root.after(0, self.finish_model_mapping)
    
    def finish_model_mapping(self):
        """Clean up after model mapping completes"""
        self.fill_mapping_button.configure(state='normal')
        self.hide_progress()
        self.set_status("Ready")

    def parse_ddl_file(self, ddl_path):
        """Parse DDL file to extract column names, data types, and default values"""
        try:
            with open(ddl_path, 'r') as file:
                ddl_content = file.read()
            
            logging.info(f"Parsing DDL file: {ddl_path}")

            # Find the create table statement
            create_table_pattern = r'CREATE\s+TABLE\s+.*?\((.*?)\)[^)]*$'
            match = re.search(create_table_pattern, ddl_content, re.DOTALL | re.IGNORECASE)
            
            if not match:
                raise Exception("Could not find CREATE TABLE statement in DDL file")
            
            columns_text = match.group(1)
            # Split by commas, but be careful with commas inside parentheses (for functions, etc.)
            column_definitions = []
            current_def = ""
            paren_level = 0
            
            for char in columns_text:
                if char == '(':
                    paren_level += 1
                    current_def += char
                elif char == ')':
                    paren_level -= 1
                    current_def += char
                elif char == ',' and paren_level == 0:
                    column_definitions.append(current_def.strip())
                    current_def = ""
                else:
                    current_def += char
                    
            # Add the last definition if it exists
            if current_def.strip():
                column_definitions.append(current_def.strip())
                
            columns = []
            unique_keys = []
            primary_keys = []

            # First pass: extract all column names
            all_column_names = []
            for definition in column_definitions:
                if not definition:
                    continue
                    
                # Skip constraints for now
                if re.match(r'^\s*(PRIMARY|FOREIGN|CONSTRAINT|UNIQUE|CHECK|INDEX)', definition, re.IGNORECASE):
                    continue
                
                # Extract column name - the first word before any whitespace
                match = re.match(r'^\s*([^\s\(]+)', definition)
                if match:
                    column_name = match.group(1).strip('"').strip('`')
                    all_column_names.append(column_name)
            
            # Second pass: process columns and constraints
            for definition in column_definitions:
                if not definition:
                    continue
                
                # Log the definition for debugging
                logging.debug(f"Processing definition: {definition}")
                    
                # Handle primary key constraints
                if re.match(r'^\s*PRIMARY\s+KEY', definition, re.IGNORECASE):
                    # Extract primary key columns
                    pk_pattern = r'PRIMARY\s+KEY\s*\(([^)]+)\)'
                    pk_match = re.search(pk_pattern, definition, re.IGNORECASE)
                    if pk_match:
                        pk_columns = [col.strip().strip('"').strip('`') for col in pk_match.group(1).split(',')]
                        # Add primary key columns to primary keys list
                        primary_keys.extend(pk_columns)
                        logging.info(f"Found PRIMARY KEY: {pk_columns}")
                    continue
                
                # Handle standalone UNIQUE constraints (not attached to a column)
                if re.match(r'^\s*UNIQUE', definition, re.IGNORECASE):
                    # Extract unique key columns
                    unique_pattern = r'UNIQUE\s*\(([^)]+)\)'
                    unique_match = re.search(unique_pattern, definition, re.IGNORECASE)
                    if unique_match:
                        key_columns = [col.strip().strip('"').strip('`') for col in unique_match.group(1).split(',')]
                        unique_keys.extend(key_columns)
                        logging.info(f"Found standalone UNIQUE constraint: {key_columns}")
                    continue
                    
                # Handle unique constraints with CONSTRAINT keyword
                if re.match(r'^\s*CONSTRAINT', definition, re.IGNORECASE) and 'UNIQUE' in definition.upper():
                    # Extract unique key columns
                    unique_key_pattern = r'CONSTRAINT\s+\w+\s+UNIQUE\s+(?:KEY\s+)?\(([^)]+)\)'
                    unique_match = re.search(unique_key_pattern, definition, re.IGNORECASE)
                    if unique_match:
                        key_columns = [col.strip().strip('"').strip('`') for col in unique_match.group(1).split(',')]
                        unique_keys.extend(key_columns)
                        logging.info(f"Found CONSTRAINT UNIQUE: {key_columns}")
                    continue
                
                # Skip other constraints
                if re.match(r'^\s*CONSTRAINT|CHECK|INDEX|FOREIGN', definition, re.IGNORECASE):
                    continue
                    
                # Skip default timestamp values or other non-column entries
                if 'CAST(CURRENT_TIMESTAMP' in definition.upper() or not definition.strip():
                    continue
                
                # Extract column name and rest of definition
                # The column name is the first word before any whitespace
                match = re.match(r'^\s*([^\s\(]+)(.*)', definition)
                if match:
                    column_name = match.group(1).strip('"').strip('`')
                    full_definition = match.group(2).strip()
                    
                    # Check if this column has a UNIQUE constraint
                    if ' UNIQUE' in full_definition.upper():
                        unique_keys.append(column_name)
                        logging.info(f"Found column with UNIQUE constraint: {column_name}")
                        # Remove the UNIQUE keyword from the definition to avoid confusion
                        full_definition = full_definition.replace(' UNIQUE', '').replace(' unique', '')
                    
                    # Add to columns list
                    columns.append((column_name, full_definition))
            
            # Look for unique index definitions outside the CREATE TABLE statement
            unique_index_pattern = r'CREATE\s+UNIQUE\s+INDEX\s+\w+\s+ON\s+\w+\s*\(([^)]+)\)'
            for unique_index_match in re.finditer(unique_index_pattern, ddl_content, re.IGNORECASE):
                index_columns = [col.strip().strip('"').strip('`') for col in unique_index_match.group(1).split(',')]
                unique_keys.extend(index_columns)
                logging.info(f"Found UNIQUE INDEX: {index_columns}")
            
            # Remove duplicates from unique keys and primary keys
            unique_keys = list(dict.fromkeys(unique_keys))
            primary_keys = list(dict.fromkeys(primary_keys))
            
            # Verify that all keys are valid column names
            valid_unique_keys = [key for key in unique_keys if key in all_column_names]
            valid_primary_keys = [key for key in primary_keys if key in all_column_names]
            
            # Store both unique and primary keys for later use
            self.unique_keys = valid_unique_keys
            self.primary_keys = valid_primary_keys
            
            logging.info(f"Parsed DDL file: Found {len(columns)} columns, {len(valid_unique_keys)} unique keys, and {len(valid_primary_keys)} primary keys")
            logging.info(f"Unique keys: {valid_unique_keys}")
            logging.info(f"Primary keys: {valid_primary_keys}")
            
            return columns, valid_unique_keys

        except Exception as e:
            logging.error(f"Error parsing DDL file: {str(e)}")
            raise Exception(f"Error parsing DDL file: {str(e)}")

    def update_mapping_sheet(self, columns):
        """Update the mapping sheet in the Excel workbook with the specified format"""
        try:
            # Create or load workbook
            if not os.path.exists(self.mapping_file_path.get()):
                workbook = openpyxl.Workbook()
                # Remove default sheet
                default_sheet = workbook['Sheet']
                workbook.remove(default_sheet)
            else:
                try:
                    workbook = load_workbook(self.mapping_file_path.get())
                    # Remove existing sheets if replacing file
                    if 'Mapping' in workbook.sheetnames:
                        workbook.remove(workbook['Mapping'])
                    if 'Config' in workbook.sheetnames:
                        workbook.remove(workbook['Config'])
                except Exception:
                    workbook = openpyxl.Workbook()
                    # Remove default sheet
                    default_sheet = workbook['Sheet']
                    workbook.remove(default_sheet)
            
            # Create Mapping and Config sheets
            mapping_sheet = workbook.create_sheet('Mapping')
            config_sheet = workbook.create_sheet('Config')
            workbook.active = mapping_sheet

            # Update sheets
            self.update_mapping_content(mapping_sheet, columns[0] if isinstance(columns, tuple) else columns)
            self.update_config_sheet(config_sheet)

            # Ensure directory exists and save
            os.makedirs(os.path.dirname(os.path.abspath(self.mapping_file_path.get())), exist_ok=True)
            try:
                workbook.save(self.mapping_file_path.get())
                messagebox.showinfo("Success", f"Mapping file created:\n{self.mapping_file_path.get()}")
            except PermissionError:
                raise Exception("Cannot save Excel file. Please close it if it's open.")
            
        except Exception as e:
            raise Exception(f"Error updating Excel file: {str(e)}")

    def update_mapping_content(self, mapping_sheet, columns):
        """Update the content of mapping sheet"""
        # Extract table name from DDL
        table_name = self.extract_table_name(self.ddl_file_path.get())
        
        # Add table name + "Mapping" as header
        mapping_sheet.merge_cells('A1:D1')  # Changed back to D1 (4 columns)
        header_cell = mapping_sheet.cell(row=1, column=1, value=f"{table_name} Mapping")
        header_cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        header_cell.fill = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Add Target and Source table rows with proper formatting
        target_label = mapping_sheet.cell(row=2, column=1, value='TARGET_TABLE')
        target_label.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        target_value = mapping_sheet.cell(row=2, column=2, value=table_name)
        
        source_label = mapping_sheet.cell(row=3, column=1, value='SOURCE_TABLE')
        source_label.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Add Source Type row
        source_type_label = mapping_sheet.cell(row=4, column=1, value='SOURCE_TYPE')
        source_type_label.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        source_type_cell = mapping_sheet.cell(row=4, column=2)
        
        # Add data validation for Source Type
        dv = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1='"source,ref"',
            allow_blank=True
        )
        mapping_sheet.add_data_validation(dv)
        dv.add(source_type_cell)

        # Add Source Name row
        source_name_label = mapping_sheet.cell(row=5, column=1, value='SOURCE_NAME')
        source_name_label.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add Materialization row
        materialization_label = mapping_sheet.cell(row=6, column=1, value='MATERIALIZATION')
        materialization_label.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        materialization_cell = mapping_sheet.cell(row=6, column=2)
        
        # Add data validation for Materialization
        mat_dv = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1='"incremental,truncate_load,lnd_load"',  # Add an empty option
            allow_blank=True
        )
        mapping_sheet.add_data_validation(mat_dv)
        mat_dv.add(materialization_cell)
        
        # Add UNIQUE_KEY row for incremental models
        unique_key_label = mapping_sheet.cell(row=7, column=1, value='UNIQUE_KEY')
        unique_key_label.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        unique_key_cell = mapping_sheet.cell(row=7, column=2)
        # Add a comment to explain the format
        comment = openpyxl.comments.Comment('For incremental models, specify comma-separated column names to use as unique keys', 'System')
        unique_key_cell.comment = comment
        
        # Add MINUS_LOGIC_REQUIRED row
        minus_logic_label = mapping_sheet.cell(row=8, column=1, value='MINUS_LOGIC_REQUIRED')
        minus_logic_label.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        minus_logic_cell = mapping_sheet.cell(row=8, column=2)
        # Add data validation for Y/N
        minus_logic_dv = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1='"Y,N"',
            allow_blank=True
        )
        mapping_sheet.add_data_validation(minus_logic_dv)
        minus_logic_dv.add(minus_logic_cell)
        # Default to N
        minus_logic_cell.value = "N"
        # Add a comment to explain the format
        minus_comment = openpyxl.comments.Comment('Set to Y to exclude audit columns and unique key combination in the minus logic', 'System')
        minus_logic_cell.comment = minus_comment
        transient_cell_label = mapping_sheet.cell(row=9, column=1, value='TRANSIENT_TABLE')
        transient_cell_label.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00",fill_type="solid")
        transient_cell = mapping_sheet.cell(row=9, column=2)
        # Add data validation for Y/N
        transient_cell_dv = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1='"Y,N"',
            allow_blank=True
        )
        mapping_sheet.add_data_validation(transient_cell_dv)
        transient_cell_dv.add(transient_cell)
        # Default to N
        transient_cell.value = "N"
        # Add a comment to explain the format
        transient_comment = openpyxl.comments.Comment('Set to Y to set true tag for transient model property', 'System')
        transient_cell.comment = transient_comment
        # Add blank row before column mappings
        mapping_sheet.append([])

        # Add column headers (start from row 10 now to account for the MINUS_LOGIC_REQUIRED row)
        headers = ['S.NO', 'TargetColumn', 'Source Table', 'Logic/Mapping/Constant Value']  # Removed Source Type and Source Name
        header_row = 11
        for col, header in enumerate(headers, start=1):
            cell = mapping_sheet.cell(row=header_row, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        # Add the data - Filter out any non-column entries like constraints
        valid_columns = []
        for column_name, definition in columns:
            # Skip if column_name is "UNIQUE", "PRIMARY", "CONSTRAINT", or any other non-column entry
            if (column_name.upper() in ["UNIQUE", "PRIMARY", "CONSTRAINT"] or 
                column_name.isdigit() or  # Skip numeric entries
                column_name.endswith(')') or  # Skip entries ending with )
                not column_name.strip()):  # Skip empty entries
                continue
                
            valid_columns.append((column_name, definition))
        
        # Now add the valid columns to the sheet
        for i, (column_name, _) in enumerate(valid_columns, start=1):
            row = header_row + i
            mapping_sheet.cell(row=row, column=1, value=i)  # S.NO
            mapping_sheet.cell(row=row, column=2, value=column_name)  # TargetColumn

        # Calculate the last row of column mappings
        last_column_row = header_row + len(valid_columns)
        
        # Add a blank row after column mappings
        last_column_row += 2
        
        # Add JOIN_TABLES section
        join_header_row = last_column_row
        join_cell = mapping_sheet.cell(row=join_header_row, column=1, value='JOIN_TABLES')
        join_cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add join table headers
        join_headers = ['Join Type', 'Table Type', 'Source Name', 'Table Name', 'Alias', 'Join Condition']
        join_header_row += 1
        for col, header in enumerate(join_headers, start=1):
            cell = mapping_sheet.cell(row=join_header_row, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        # Add 5 empty rows for joins
        for i in range(5):
            join_header_row += 1
            row = join_header_row
            
            # Add default value for Join Type (LEFT)
            join_type_cell = mapping_sheet.cell(row=row, column=1, value='LEFT')
            
            # Add data validation for Join Type
            join_type_dv = openpyxl.worksheet.datavalidation.DataValidation(
                type="list",
                formula1='"LEFT,INNER"',
                allow_blank=True
            )
            mapping_sheet.add_data_validation(join_type_dv)
            join_type_dv.add(join_type_cell)
            
            # Add data validation for Table Type
            table_type_cell = mapping_sheet.cell(row=row, column=2)
            table_type_dv = openpyxl.worksheet.datavalidation.DataValidation(
                type="list",
                formula1='"source,ref"',
                allow_blank=True
            )
            mapping_sheet.add_data_validation(table_type_dv)
            table_type_dv.add(table_type_cell)
        
        # Add WHERE_CONDITIONS section
        where_row = join_header_row + 2
        where_cell = mapping_sheet.cell(row=where_row, column=1, value='WHERE_CONDITIONS')
        where_cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add where condition input cell with more space
        where_input_cell = mapping_sheet.cell(row=where_row, column=2)
        # Merge cells for more space
        mapping_sheet.merge_cells(start_row=where_row, start_column=2, end_row=where_row, end_column=6)
        
        # Add GROUP_BY section
        group_row = where_row + 2
        group_cell = mapping_sheet.cell(row=group_row, column=1, value='GROUP BY')
        group_cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add group by input cell with more space
        group_input_cell = mapping_sheet.cell(row=group_row, column=2)
        # Merge cells for more space
        mapping_sheet.merge_cells(start_row=group_row, start_column=2, end_row=group_row, end_column=6)
        
        # Adjust column widths
        column_widths = {
            'A': 15,   # S.NO/Join Type
            'B': 40,   # TargetColumn/Table Type
            'C': 40,   # Source Table/Source Name
            'D': 50,   # Logic/Table Name
            'E': 20,   # Alias
            'F': 60,   # Join Condition
        }
        
        for col, width in column_widths.items():
            mapping_sheet.column_dimensions[col].width = width

    def update_config_sheet(self, config_sheet):
        """Create and update the Config sheet"""
        # Add header
        config_sheet.merge_cells('A1:C1')
        header_cell = config_sheet.cell(row=1, column=1, value='DAG Configuration')
        header_cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        header_cell.fill = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Add DAG Configuration section
        config_sheet.cell(row=2, column=1, value='DAG Configuration').font = openpyxl.styles.Font(bold=True)
        
        # Update DAG config options
        dag_options = [
            {
                'param': 'DAG Type',
                'value': '',
                'options': ['DATASET DEPENDENCY', 'CRON', 'SNS'],
                'description': 'Select the type of DAG to generate'
            },
            {
                'param': 'Schedule Interval',
                'value': '0 */4 * * *',
                'description': 'Cron expression for scheduling (e.g., 0 */4 * * * for every 4 hours)'
            }
        ]

        # Add DAG config rows
        current_row = 3
        for config in dag_options:
            # Parameter name with yellow background
            param_cell = config_sheet.cell(row=current_row, column=1, value=config['param'])
            param_cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Value cell
            value_cell = config_sheet.cell(row=current_row, column=2)
            value_cell.value = config['value']
            
            # Description
            desc_cell = config_sheet.cell(row=current_row, column=3, value=config['description'])
            
            # If it's DAG Type, create a data validation dropdown
            if config['param'] == 'DAG Type':
                dv = openpyxl.worksheet.datavalidation.DataValidation(
                    type="list",
                    formula1=f'"{",".join(config["options"])}"',
                    allow_blank=True
                )
                config_sheet.add_data_validation(dv)
                dv.add(value_cell)
            
            current_row += 1

        # Add Dependencies section
        current_row += 1
        dep_header = config_sheet.cell(row=current_row, column=1, value='Dependencies')
        dep_header.font = openpyxl.styles.Font(bold=True)
        dep_header.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        current_row += 1

        # Add dependency column headers
        headers = ['#', 'Dependency Schema', 'Dependency Table']
        for col, header in enumerate(headers, start=1):
            cell = config_sheet.cell(row=current_row, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        # Add 5 empty rows for dependencies
        for i in range(5):
            current_row += 1
            row_num = config_sheet.cell(row=current_row, column=1, value=i + 1)
            row_num.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Add example values for the first row
            if i == 0:
                config_sheet.cell(row=current_row, column=2, value="DW")
                config_sheet.cell(row=current_row, column=3, value="D_CUSTOMER")
                
                # Add comment to explain the format
                comment = openpyxl.comments.Comment('Enter the schema and table name of the dependency', 'System')
                config_sheet.cell(row=current_row, column=2).comment = comment

        # Add blank row for separation
        current_row += 2

        # Add Snowflake Configuration section
        config_sheet.cell(row=current_row, column=1, value='Snowflake Configuration').font = openpyxl.styles.Font(bold=True)
        current_row += 1

        # Snowflake config options
        snowflake_options = [
            {
                'param': 'ROLE',
                'value': 'REV_GROWTH_MGMT_DEV_EDITOR_ROLE',
                'description': 'Snowflake role to use'
            },
            {
                'param': 'WAREHOUSE',
                'value': 'REV_GROWTH_MGMT_DEV_LOAD_WH',
                'description': 'Snowflake warehouse to use'
            },
            {
                'param': 'DATABASE',
                'value': 'REV_GROWTH_MGMT_DEV_DB',
                'description': 'Snowflake database to use'
            },
            {
                'param': 'ACCOUNT',
                'value': 'cbrands.us-east-1',
                'description': 'Snowflake account identifier'
            },
            {
                'param': 'USER',
                'value': '',
                'description': 'Snowflake username'
            },
            {
                'param': 'AUTHENTICATOR',
                'value': 'externalbrowser',
                'description': 'Authentication method'
            }
        ]

        # Add Snowflake config rows
        for config in snowflake_options:
            # Parameter name with yellow background
            param_cell = config_sheet.cell(row=current_row, column=1, value=config['param'])
            param_cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Value cell
            value_cell = config_sheet.cell(row=current_row, column=2, value=config['value'])
            
            # Description
            desc_cell = config_sheet.cell(row=current_row, column=3, value=config['description'])
            
            current_row += 1

        # Adjust column widths
        config_sheet.column_dimensions['A'].width = 20  # Parameter
        config_sheet.column_dimensions['B'].width = 40  # Value
        config_sheet.column_dimensions['C'].width = 60  # Description

        # Add borders to all cells
        for row in config_sheet.iter_rows(min_row=2, max_row=current_row):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

        # Add section headers styling
        for row in [2, current_row - len(snowflake_options) - 1]:
            cell = config_sheet.cell(row=row, column=1)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    def extract_table_name(self, ddl_path):
        """Extract table name from DDL file"""
        try:
            with open(ddl_path, 'r') as file:
                ddl_content = file.read()
            
            # Find the table name in CREATE TABLE statement
            # This pattern will handle:
            # - IF NOT EXISTS clause
            # - Schema prefix
            # - Quoted and unquoted identifiers
            table_pattern = r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(?:\"?(\w+)\"?\.)?"?(\w+)"?'
            match = re.search(table_pattern, ddl_content, re.IGNORECASE)
            
            if match:
                # Get schema and table name
                schema = match.group(1) or ''  # First group (schema) - optional
                table = match.group(2)         # Second group (table name)
                
                # Remove any quotes and trim
                schema = schema.strip('"').strip('`').strip()
                table = table.strip('"').strip('`').strip()
                
                # Combine schema and table if schema exists
                full_table_name = f"{schema}.{table}" if schema else table
                return full_table_name.upper()  # Convert to uppercase
            
            return "Unknown_Table"
        except Exception:
            return "Unknown_Table"

    def show_error(self, message):
        messagebox.showerror("Error", message)
        self.progress.stop()
        self.progress.pack_forget()  # Changed from grid_remove to pack_forget
        self.root.after(350, self.root.destroy)  # Close the app within a seconds

    def on_tab_change(self, event):
        """Handle tab change with smooth animation"""
        try:
            current_tab = self.notebook.select()
            if not current_tab:
                return
            
            # Get the currently selected tab widget
            tab_frame = self.notebook.children[current_tab.split('.')[-1]]
            
            # Start fade-in animation
            self.animate_frame(tab_frame)
            
        except Exception as e:
            print(f"Animation error: {str(e)}")

    def animate_frame(self, frame, step=0):
        """Smooth frame animation"""
        try:
            if step <= self.animation_steps:
                # Calculate opacity
                opacity = step / self.animation_steps
                
                # Apply the fade effect
                frame.update()
                
                if step < self.animation_steps:
                    # Schedule next animation step
                    delay = self.animation_duration // self.animation_steps
                    self.root.after(delay, lambda: self.animate_frame(frame, step + 1))
                
        except Exception as e:
            print(f"Animation step error: {str(e)}")

    def add_to_history(self, file_path, history_list):
        """Add file to history and maintain uniqueness"""
        if file_path in history_list:
            history_list.remove(file_path)
        history_list.insert(0, file_path)
        del history_list[10:]  # Keep only last 10 items
        self.save_history()

    def save_history(self):
        """Save file history to a JSON file"""
        try:
            history = {
                'mapping_files': self.file_history[-10:],  # Keep last 10 files
                'ddl_files': self.ddl_file_history[-10:]   # Keep last 10 files
            }
            history_file = os.path.join(os.getcwd(), 'file_history.json')
            with open(history_file, 'w') as f:
                json.dump(history, f, indent=4)
        except Exception as e:
            print(f"Error saving history: {str(e)}")

    def update_ui_with_resources(self):
        """Update UI components with loaded resources"""
        try:
            # Update combobox values with formatted paths
            if hasattr(self, 'file_combo'):
                formatted_values = [self.format_path(path) for path in self.file_history]
                self.file_combo['values'] = formatted_values
            if hasattr(self, 'ddl_combo'):
                formatted_values = [self.format_path(path) for path in self.ddl_file_history]
                self.ddl_combo['values'] = formatted_values
            if hasattr(self, 'mapping_combo'):
                formatted_values = [self.format_path(path) for path in self.file_history]
                self.mapping_combo['values'] = formatted_values
            
            self.root.update_idletasks()
        except Exception as e:
            logging.error(f"Error updating UI: {str(e)}")

    def on_combo_select(self, event):
        """Handle combobox selection"""
        combo = event.widget
        selected_display = combo.get()
        
        # Find the full path from the display value
        for full_path in self.file_history:
            if self.format_path(full_path) == selected_display:
                self.mapping_file_path.set(full_path)
                break

    def format_path(self, path):
        """Format path for display in combobox"""
        if not path:
            return ''
        # Show filename and parent directory
        try:
            dirname, filename = os.path.split(path)
            parent_dir = os.path.basename(dirname)
            return f"{parent_dir}/{filename}"
        except:
            return path

    def create_mapping_template(self, file_path):
        """Create a mapping template Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapping"

        # Add headers and formatting
        headers = [
            ("TARGET_TABLE", ""),
            ("SOURCE_TABLE", ""),
            ("SOURCE_TYPE", "source"),
            ("SOURCE_NAME", ""),
            ("MATERIALIZATION", "incremental"),
            ("UNIQUE_KEY", ""),
            ("MINUS_LOGIC_REQUIRED", "N"),
            ("MERGE_UPDATE_EXCLUDE_COLUMNS", "CREATE_DT,CREATE_BY"),
            ("", "")
        ]

        # Apply headers
        for i, (header, default) in enumerate(headers, 1):
            ws.cell(row=i, column=1, value=header)
            ws.cell(row=i, column=2, value=default)
            ws.cell(row=i, column=1).font = Font(bold=True)

if __name__ == "__main__":
    # Initialize logging first
    setup_logging()
    
    # Save original stdout for printing themes
    original_stdout = sys.stdout
    
    try:
        root = ThemedTk()
        
        # # Temporarily restore stdout to print themes to console
        # sys.stdout = original_stdout
        # print("Available themes:", root.get_themes())
        #
        # # Redirect back to logger
        # if logging.getLogger().handlers:
        #     sys.stdout = logging.getLogger().handlers[0].stream
        #
        # # Log themes to log file as well
        # logging.info(f"Available themes: {root.get_themes()}")
        #
        # root.set_theme("breeze")
        # logging.info("Using theme: breeze")
        
        app = DAGGeneratorApp(root)
        root.mainloop()
    except Exception as e:
        logging.error(f"Application error: {str(e)}")
        logging.error(f"Traceback: {traceback.format_exc()}")
        if original_stdout:
            sys.stdout = original_stdout
            print(f"Error: {str(e)}")
            print("Check app.log for details")
